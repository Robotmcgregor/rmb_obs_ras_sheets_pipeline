#####
# This script takes the CSV field data output from ODK Briefcase and outputs an Excel spreadsheet for each integrated site,
# which can include basal and woody thickening. Additional outputs are a summary Excel spreadsheet and shapefile for all sites,
# and a spreadhseet of photos for each site.
# Author: Dan McIntyre
#####

import xlsxwriter
from openpyxl import load_workbook
import pandas as pd
import geopandas
from shapely.geometry import Point

pd.options.mode.chained_assignment = None  # Gets rid of the pesky 'SettingWithCopyWarning'...


def SplitKey(transect1_point, transect2_point, transect3_point):
    cover_df1 = transect1_point[['KEY', 'REPEAT_COUNT1', 'GROUND1', 'BELOW1', 'ABOVE1', 'SET-OF-REPEAT_points_1']]
    cover_df1.columns = (['KEY', 'REPEAT_COUNT', 'GROUND', 'BELOW', 'ABOVE', 'PARENT_KEY'])

    cover_df2 = transect2_point[['KEY', 'REPEAT_COUNT2', 'GROUND2', 'BELOW2', 'ABOVE2', 'SET-OF-REPEAT_points_2']]
    cover_df2.columns = (['KEY', 'REPEAT_COUNT', 'GROUND', 'BELOW', 'ABOVE', 'PARENT_KEY'])

    cover_df3 = transect3_point[['KEY', 'REPEAT_COUNT3', 'GROUND3', 'BELOW3', 'ABOVE3', 'SET-OF-REPEAT_points_3']]
    cover_df3.columns = (['KEY', 'REPEAT_COUNT', 'GROUND', 'BELOW', 'ABOVE', 'PARENT_KEY'])

    transect_point = pd.concat([cover_df1, cover_df2, cover_df3], axis=0)

    # Split the 'Key' name so that multiple data frames can be merged.
    key_split = transect_point["KEY"].str.split("/", expand=True)
    key_split.columns = ["KEY", "1"]
    key_split = key_split["KEY"]

    # Extract the stripped key
    transect_point["NEW_KEY"] = key_split

    cover_df = transect_point[['NEW_KEY', 'REPEAT_COUNT', 'GROUND', 'BELOW', 'ABOVE', 'PARENT_KEY']]
    cover_df.columns = (['KEY', 'REPEAT_COUNT', 'GROUND', 'BELOW', 'ABOVE', 'PARENT_KEY'])

    return cover_df


def PrepCover(transect_basic, cover_df):
    # Extract the site name from the csv
    key_extract = transect_basic.filter(['SITE', 'KEY'], axis=1)

    # Merge the two dataframes
    merged_df1 = pd.merge(key_extract, cover_df, on='KEY')

    # Extract the transect parent key and name from 'STAR_TRANSECT.csv'
    name_extract1 = transect_basic.filter(['TRAN1', 'SET-OF-REPEAT_points_1'])
    name_extract1.columns = (['TRAN', 'PARENT_KEY'])
    name_extract2 = transect_basic.filter(['TRAN2', 'SET-OF-REPEAT_points_2'])
    name_extract2.columns = (['TRAN', 'PARENT_KEY'])
    name_extract3 = transect_basic.filter(['TRAN3', 'SET-OF-REPEAT_points_3'])
    name_extract3.columns = (['TRAN', 'PARENT_KEY'])

    # Combine the three separate dfs into one df
    name_extract = pd.concat([name_extract1, name_extract2, name_extract3], axis=0)

    # Merge the two dataframes
    merged_df2 = pd.merge(name_extract, merged_df1, on="PARENT_KEY")

    final_cover_df = merged_df2[['SITE', 'TRAN', 'REPEAT_COUNT', 'GROUND', 'BELOW', 'ABOVE']]

    # Change values in columns to required values for workbook
    ground_values = {'bare': 'BARE GROUND', 'gravel': 'GRAVEL', 'rock': 'ROCK', 'ash': 'ASH', 'litter': 'LITTER',
                     'cryptogram': 'CRYPTOGRAM',
                     'dead_annual_grass': 'DEAD ANNUAL GRASS', 'dead_perennial_grass': 'DEAD PERENNIAL GRASS',
                     'dead_annual_forb':
                         'DEAD ANNUAL FORB / HERB', 'dead_perennial_forb': 'DEAD PERENNIAL FORB / HERB',
                     'green_annual_grass': 'GREEN ANNUAL GRASS',
                     'green_perennial_grass': 'GREEN PERENNIAL GRASS', 'green_annual_forb': 'GREEN ANNUAL FORB / HERB',
                     'green_perennial_forb':
                         'GREEN PERENNIAL FORB / HERB', 'green_plant': 'GREEN PLANT', 'dead_plant': 'DEAD PLANT'}

    below_values = {'below_green': 'BELOW - GREEN', 'below_brown': 'BELOW - BROWN', 'below_dead': 'BELOW - DEAD',
                    'subshrub': 'SUBSHRUB - GREY', 'none': 'BLANK'}

    above_values = {'above_green': 'ABOVE - GREEN', 'above_brown': 'ABOVE - BROWN', 'above_dead': 'ABOVE - DEAD',
                    'above_in_crown':
                        'ABOVE - IN CROWN', 'not_in_crown': 'BLANK'}

    tran_values = {'north_south': 'NORTH', 'southeast_northwest': 'SOUTH EAST', 'southwest_northeast': 'SOUTH WEST'}

    final_cover_df['GROUND'] = final_cover_df['GROUND'].replace(ground_values)
    final_cover_df['BELOW'] = final_cover_df['BELOW'].replace(below_values)
    final_cover_df['ABOVE'] = final_cover_df['ABOVE'].replace(above_values)
    final_cover_df['TRAN'] = final_cover_df['TRAN'].replace(tran_values)

    return final_cover_df


def CreateWorkbook(path, site):
    """Recreate the Rangeland Montioring Excel workbook with the xlsxwriter module"""

    workbook = xlsxwriter.Workbook(path + site + '.xlsx')

    # Define format variables

    Heading1 = workbook.add_format()
    Heading1.set_font_name('Calibri')
    Heading1.set_font_size(32)
    Heading1.set_bold()
    Heading1.set_align('center')
    Heading1.set_align('vcenter')
    Heading1.set_text_wrap()
    Heading1.set_bg_color('#fcbd00')
    Heading1.set_border()
    Heading1.set_border_color('#C0C0C0')

    Heading2 = workbook.add_format()
    Heading2.set_font_name('Calibri')
    Heading2.set_font_size(20)
    Heading2.set_bold()
    Heading2.set_align('center')
    Heading2.set_align('vcenter')
    Heading2.set_text_wrap()
    Heading2.set_bg_color('#fcbd00')
    Heading2.set_border()
    Heading2.set_border_color('#C0C0C0')

    Heading3 = workbook.add_format()
    Heading3.set_font_name('Calibri')
    Heading3.set_font_size(16)
    Heading3.set_bold()
    Heading3.set_align('right')
    Heading3.set_align('vcenter')
    Heading3.set_text_wrap()
    Heading3.set_bg_color('#fcbd00')
    Heading3.set_border()
    Heading3.set_border_color('#C0C0C0')

    # Same as Heading3, except center aligned instead of right

    Heading4 = workbook.add_format()
    Heading4.set_font_name('Calibri')
    Heading4.set_font_size(16)
    Heading4.set_bold()
    Heading4.set_align('center')
    Heading4.set_align('vcenter')
    Heading4.set_text_wrap()
    Heading4.set_bg_color('#fcbd00')
    Heading4.set_border()
    Heading4.set_border_color('#C0C0C0')

    # Smaller font used for items

    Heading5 = workbook.add_format()
    Heading5.set_font_name('Calibri')
    Heading5.set_font_size(14)
    Heading5.set_bold()
    Heading5.set_align('center')
    Heading5.set_align('vcenter')
    Heading5.set_text_wrap()
    Heading5.set_bg_color('#fedf98')
    Heading5.set_border()
    Heading5.set_border_color('#C0C0C0')

    # Smaller font used for uncolored items

    Heading6 = workbook.add_format()
    Heading6.set_font_name('Calibri')
    Heading6.set_font_size(14)
    Heading6.set_bold()
    Heading6.set_align('center')
    Heading6.set_align('vcenter')
    Heading6.set_text_wrap()
    Heading6.set_border()
    Heading6.set_border_color('#C0C0C0')

    # Font used for data values

    Heading7 = workbook.add_format()
    Heading7.set_font_name('Calibri')
    Heading7.set_font_size(14)
    Heading7.set_align('center')
    Heading7.set_align('vcenter')
    Heading7.set_text_wrap()
    Heading7.set_border()
    Heading7.set_border_color('#C0C0C0')

    # Font used for data values w. italics
    Heading8 = workbook.add_format()
    Heading8.set_font_name('Calibri')
    Heading8.set_font_size(14)
    Heading8.set_italic()
    Heading8.set_align('center')
    Heading8.set_align('vcenter')
    Heading8.set_text_wrap()
    Heading8.set_border()
    Heading8.set_border_color('#C0C0C0')

    # Used to fill in emply boxes with color

    ColorFill = workbook.add_format()
    ColorFill.set_bg_color('#fcbd00')
    ColorFill.set_border()
    ColorFill.set_border_color('#C0C0C0')

    # Set up Site Establishment worksheet #

    worksheet1 = workbook.add_worksheet('Step 1 - Site Establishment')

    # Set row height

    worksheet1.set_default_row(56.25)

    # Add item headings to cells as strings

    worksheet1.write_string('A2', 'ITEM', Heading2)
    worksheet1.write_string('A3', 'RECORDER:', Heading3)
    worksheet1.write_string('A4', 'ESTIMATOR:', Heading3)
    worksheet1.write_string('A5', 'ANY OTHERS PRESENT:', Heading3)
    worksheet1.write_string('A6', 'PROPERTY:', Heading3)
    worksheet1.write_string('A7', 'UNLISTED PROPERTY: (if required)', Heading3)
    worksheet1.write_string('A8', 'SITE ID:', Heading3)
    worksheet1.write_string('A9', 'PADDOCK NAME:', Heading3)
    worksheet1.write_string('A10', 'DATE & TIME:', Heading3)
    worksheet1.write_string('A11', 'DIRECTION FROM OFFSET:', Heading3)
    worksheet1.write_string('A12', 'DATUM:', Heading3)
    worksheet1.write_string('A13', 'OFFSET (LAT):', Heading3)
    worksheet1.write_string('A14', 'OFFSET (LONG):', Heading3)
    worksheet1.write_string('A15', 'CENTRE POINT (LAT):', Heading3)
    worksheet1.write_string('A16', 'CENTRE POINT (LONG):', Heading3)
    worksheet1.write_string('A17', 'LANDSCAPE POSITION:', Heading3)
    worksheet1.write_string('A18', 'SOIL SURFACE COLOUR:', Heading3)
    worksheet1.write_string('A19', 'SITE DESCRIPTION:', Heading3)
    worksheet1.write_string('A20', 'REASON FOR SITE SELECTION:', Heading3)
    worksheet1.write_string('A21', 'LAND SYSTEM:', Heading3)
    worksheet1.write_string('D21', 'SOURCE', Heading3)
    worksheet1.write_string('A22', 'CONSISTENT WITH MAPPING?:', Heading3)
    worksheet1.write_string('D22', 'IF NO, WHAT LAND SYSTEM:', Heading3)
    worksheet1.write_string('A23', 'NOTES:', Heading3)
    worksheet1.write_string('A24', 'NEAREST STOCK WATER (NAME):', Heading3)
    worksheet1.write_string('D24', 'NEAREST STOCK WATER (TYPE):', Heading3)
    worksheet1.write_string('A25', 'DISTANCE FROM NEAREST STOCK WATER (km):', Heading3)
    worksheet1.write_string('D25', 'DIRECTION FROM NEAREST STOCK WATER:', Heading3)
    worksheet1.write_string('A26', 'DISTANCE FROM TRACK (m):', Heading3)
    worksheet1.write_string('D26', 'DIRECTION FROM TRACK:', Heading3)
    worksheet1.write_string('A27', 'OTHER NEARBY INFRASTRUCTURE:', Heading3)

    # Add blank cells with formatting to contain data

    worksheet1.write_blank('E21', None, Heading7)
    worksheet1.write_blank('E22', None, Heading7)
    worksheet1.write_blank('E24', None, Heading7)
    worksheet1.write_blank('E25', None, Heading7)
    worksheet1.write_blank('E26', None, Heading7)

    # Merge the relevant cells

    worksheet1.merge_range('A1:E1', 'Step 1 - SITE ESTABLISHMENT', Heading1)
    worksheet1.merge_range('B2:E2', 'INPUT', Heading2)
    worksheet1.merge_range('B3:E3', '', Heading7)
    worksheet1.merge_range('B4:E4', '', Heading7)
    worksheet1.merge_range('B5:E5', '', Heading7)
    worksheet1.merge_range('B6:E6', '', Heading7)
    worksheet1.merge_range('B7:E7', '', Heading7)
    worksheet1.merge_range('B8:E8', '', Heading7)
    worksheet1.merge_range('B9:E9', '', Heading7)
    worksheet1.merge_range('B10:E10', '', Heading7)
    worksheet1.merge_range('B11:E11', '', Heading7)
    worksheet1.merge_range('B12:E12', '', Heading7)
    worksheet1.merge_range('B13:E13', '', Heading7)
    worksheet1.merge_range('B14:E14', '', Heading7)
    worksheet1.merge_range('B15:E15', '', Heading7)
    worksheet1.merge_range('B16:E16', '', Heading7)
    worksheet1.merge_range('B17:E17', '', Heading7)
    worksheet1.merge_range('B18:E18', '', Heading7)
    worksheet1.merge_range('B19:E19', '', Heading7)
    worksheet1.merge_range('B20:E20', '', Heading7)
    worksheet1.merge_range('B21:C21', '', Heading7)
    worksheet1.merge_range('B22:C22', '', Heading7)
    worksheet1.merge_range('B23:E23', '', Heading7)
    worksheet1.merge_range('B24:C24', '', Heading7)
    worksheet1.merge_range('B25:C25', '', Heading7)
    worksheet1.merge_range('B26:C26', '', Heading7)
    worksheet1.merge_range('B27:E27', '', Heading7)

    # Set column width

    worksheet1.set_column('A:A', 45.73)
    worksheet1.set_column('B:B', 10.09)
    worksheet1.set_column('C:C', 26.91)
    worksheet1.set_column('D:D', 26.91)
    worksheet1.set_column('E:E', 26.91)

    # Setup Visit Details worksheet #

    worksheet2 = workbook.add_worksheet('Step 2 - Visit Details')

    # Set row height

    worksheet2.set_default_row(56.25)

    # Add item headings to cells as strings

    worksheet2.write_string('A2', 'ITEM', Heading2)
    worksheet2.write_string('B2', 'INPUT', Heading2)
    worksheet2.write_string('A3', 'RECORDER:', Heading3)
    worksheet2.write_string('A4', 'ESTIMATOR:', Heading3)
    worksheet2.write_string('A5', 'SITE ID:', Heading3)
    worksheet2.write_string('A6', 'DATE & TIME:', Heading3)
    worksheet2.write_string('A7', 'OFFSET PHOTO NUMBER:', Heading3)
    worksheet2.write_string('A8', 'SEASONAL CONDITIONS:', Heading3)
    worksheet2.write_string('A9', 'ATMOSPHERIC CONDITIONS:', Heading3)
    worksheet2.write_string('A10', 'SURFACE CRACKS:', Heading3)
    worksheet2.write_string('A11', 'SOIL MOISTURE:', Heading3)
    worksheet2.write_string('A12', 'BRIEF SITE DESCRIPTION:', Heading3)

    # Add blank cells with formatting to contain data

    worksheet2.write_blank('B3', None, Heading7)
    worksheet2.write_blank('B4', None, Heading7)
    worksheet2.write_blank('B5', None, Heading7)
    worksheet2.write_blank('B6', None, Heading7)
    worksheet2.write_blank('B7', None, Heading7)
    worksheet2.write_blank('B8', None, Heading7)
    worksheet2.write_blank('B9', None, Heading7)
    worksheet2.write_blank('B10', None, Heading7)
    worksheet2.write_blank('B11', None, Heading7)
    worksheet2.write_blank('B12', None, Heading7)

    # merge the relevant cells
    worksheet2.merge_range('A1:B1', 'STEP 2 - VISIT DETAILS', Heading1)

    # set column width
    worksheet2.set_column('A:A', 45.73)
    worksheet2.set_column('B:B', 71.09)

    # Set up Disturbance Details #

    worksheet3 = workbook.add_worksheet('Step 3 - Disturbance Details')

    worksheet3.set_default_row(57.00)

    # Add item headings to cells as strings

    worksheet3.write_string('A2', 'ITEM', Heading2)
    worksheet3.write_string('A3', 'CLEARING:', Heading3)
    worksheet3.write_string('A5', 'CYCLONE/STORM:', Heading3)
    worksheet3.write_string('A7', 'DIEBACK', Heading3)
    worksheet3.write_string('A9', 'ADJACENT INFRASTRUCTURE:', Heading3)
    worksheet3.write_string('A10', 'COMMENTS:', Heading3)
    worksheet3.write_string('A11', 'DISTANCE TO INFRASTRUCTURE:', Heading3)
    worksheet3.write_string('A12', 'WILD ANIMAL ACTIVITY:', Heading3)
    worksheet3.write_string('B12', 'Camel', Heading5)
    worksheet3.write_string('C12', 'Rabbit', Heading5)
    worksheet3.write_string('D12', 'Donkey', Heading5)
    worksheet3.write_string('E12', 'Horse', Heading5)
    worksheet3.write_string('F12', 'Pig', Heading5)
    worksheet3.write_string('G12', 'Buffalo', Heading5)
    worksheet3.write_string('H12', 'Native herbivore', Heading5)
    worksheet3.write_string('I12', 'Other', Heading5)
    worksheet3.write_string('A13', 'ACTIVE', Heading3)
    worksheet3.write_string('A14', 'EVIDENCE AND DESCRIPTION:', Heading3)
    worksheet3.write_string('A17', 'FREQUENCY', Heading3)
    worksheet3.write_string('A18', 'INTENSITY', Heading3)
    worksheet3.write_string('A20', 'CATTLE ACTIVITY:', Heading3)
    worksheet3.write_string('A22', 'IS THIS AN ERODIBLE SOIL?', Heading3)
    worksheet3.write_string('B23', 'Severity', Heading4)
    worksheet3.write_string('A24', 'SCALDING - wind or water', Heading3)
    worksheet3.write_string('A25', 'WINDSHEETING', Heading3)
    worksheet3.write_string('A26', 'WATERSHEETING', Heading3)
    worksheet3.write_string('A27', 'RILLING', Heading3)
    worksheet3.write_string('A28', 'GULLYING', Heading3)
    worksheet3.write_string('A29', 'EROSION COMMENTS AND PHOTO NUMBERS', Heading3)
    worksheet3.write_string('A30', 'WEEDS', Heading3)
    worksheet3.write_string('A31', 'OTHER DISTURBANCE COMMENTS', Heading3)

    worksheet3.write('A4', None, ColorFill)
    worksheet3.write('A8', None, ColorFill)
    worksheet3.write('A16', None, ColorFill)
    worksheet3.write('A19', None, ColorFill)
    worksheet3.write('A23', None, ColorFill)

    # Add blank cells with formatting to contain data

    worksheet3.write_blank('B13', None, Heading7)
    worksheet3.write_blank('C13', None, Heading7)
    worksheet3.write_blank('D13', None, Heading7)
    worksheet3.write_blank('E13', None, Heading7)
    worksheet3.write_blank('F13', None, Heading7)
    worksheet3.write_blank('G13', None, Heading7)
    worksheet3.write_blank('H13', None, Heading7)
    worksheet3.write_blank('I13', None, Heading7)
    worksheet3.write_blank('K20', None, Heading7)
    worksheet3.write_blank('B24', None, Heading7)
    worksheet3.write_blank('B25', None, Heading7)
    worksheet3.write_blank('B26', None, Heading7)
    worksheet3.write_blank('B27', None, Heading7)
    worksheet3.write_blank('B28', None, Heading7)

    # merge the relevant cells

    worksheet3.merge_range('A1:K1', 'STEP 3 - DISTURBANCE DETAILS', Heading1)
    worksheet3.merge_range('B2:K2', 'INPUT', Heading2)
    worksheet3.merge_range('B3:C3', '', Heading7)
    worksheet3.merge_range('D3:E3', 'Photo numbers', Heading4)
    worksheet3.merge_range('F3:K3', '', Heading7)
    worksheet3.merge_range('B4:C4', 'Comments', Heading4)
    worksheet3.merge_range('D4:K4', '', Heading7)
    worksheet3.merge_range('B5:C5', '', Heading7)
    worksheet3.merge_range('D5:E5', 'Photo numbers', Heading4)
    worksheet3.merge_range('F5:K5', '', Heading7)
    worksheet3.merge_range('B6:C6', 'Comments', Heading4)
    worksheet3.merge_range('D6:K6', '', Heading7)
    worksheet3.merge_range('B7:C7', '', Heading7)
    worksheet3.merge_range('D7:E7', 'Photo numbers', Heading4)
    worksheet3.merge_range('F7:K7', '', Heading7)
    worksheet3.merge_range('B8:C8', 'Comments', Heading4)
    worksheet3.merge_range('D8:K8', '', Heading7)
    worksheet3.merge_range('B9:K9', '', Heading7)
    worksheet3.merge_range('B10:K10', '', Heading7)
    worksheet3.merge_range('B11:K11', '', Heading7)
    worksheet3.merge_range('B14:K14', '', Heading7)
    worksheet3.merge_range('A15:K15', 'FIRE', Heading2)
    worksheet3.merge_range('B16:F16', 'NORTH REGION', Heading4)
    worksheet3.merge_range('B17:F17', '', Heading7)
    worksheet3.merge_range('G17:K17', '', Heading7)
    worksheet3.merge_range('B18:F18', '', Heading7)
    worksheet3.merge_range('G16:K16', 'SOUTH REGION', Heading4)
    worksheet3.merge_range('G18:K18', '', Heading7)
    worksheet3.merge_range('B19:K19', '')  # No formatting required
    worksheet3.merge_range('B20:C20', 'CATTLE PADS', Heading4)
    worksheet3.merge_range('D20:F20', '', Heading7)
    worksheet3.merge_range('G20:J20', 'TRAMPLING', Heading4)
    worksheet3.merge_range('A21:K21', 'EROSION', Heading2)
    worksheet3.merge_range('B22:K22', '', Heading7)
    worksheet3.merge_range('C23:D23', 'Stability', Heading4)
    worksheet3.merge_range('E23:K23', '', Heading7)
    worksheet3.merge_range('C24:D24', '', Heading7)
    worksheet3.merge_range('E24:K24', '', Heading7)
    worksheet3.merge_range('C25:D25', '', Heading7)
    worksheet3.merge_range('E25:K25', '', Heading7)
    worksheet3.merge_range('C26:D26', '', Heading7)
    worksheet3.merge_range('E26:K26', '', Heading7)
    worksheet3.merge_range('C27:D27', '', Heading7)
    worksheet3.merge_range('E27:K27', '', Heading7)
    worksheet3.merge_range('C28:D28', '', Heading7)
    worksheet3.merge_range('E28:K28', '', Heading7)
    worksheet3.merge_range('B29:K29', '', Heading7)
    worksheet3.merge_range('B30:K30', '', Heading7)
    worksheet3.merge_range('B31:K31', '', Heading7)

    # set column width

    worksheet3.set_column('A:A', 37.91)
    worksheet3.set_column('B:B', 12.27)
    worksheet3.set_column('C:C', 8.45)
    worksheet3.set_column('D:D', 10.82)
    worksheet3.set_column('E:E', 8.14)
    worksheet3.set_column('F:F', 8.09)
    worksheet3.set_column('G:G', 8.45)
    worksheet3.set_column('H:H', 12.14)
    worksheet3.set_column('I:I', 8.14)
    worksheet3.set_column('J:J', 6.82)
    worksheet3.set_column('K:K', 17.55)

    # Set up Transect 1 worksheet #

    worksheet4 = workbook.add_worksheet('Step 4A - Transect 1')

    worksheet4.set_default_row(56.25)

    # Add item headings to cells as strings

    x = range(1, 101)

    worksheet4.write_string('B2', 'TRANSECT 1', Heading2)
    worksheet4.write_string('C2', 'START POINT', Heading2)
    worksheet4.write_string('B3', 'GROUND LAYER', Heading4)
    worksheet4.write_string('C3', 'BELOW', Heading4)
    worksheet4.write_string('D3', 'ABOVE', Heading4)
    worksheet4.write_column('A4', x, Heading4)
    worksheet4.write('A1', None, ColorFill)
    worksheet4.write('A2', None, ColorFill)
    worksheet4.write('A3', None, ColorFill)

    # set column width

    worksheet4.set_column('A:A', 9.09)
    worksheet4.set_column('B:B', 48.00)
    worksheet4.set_column('C:C', 48.00)
    worksheet4.set_column('D:D', 48.00)
    worksheet4.set_column('E:E', 13.00)

    # Add blank cells with formatting to contain data

    worksheet4.write_column('B4', x, Heading7)
    worksheet4.write_column('C4', x, Heading7)
    worksheet4.write_column('D4', x, Heading7)

    # merge the relevant cells

    worksheet4.merge_range('B1:E1', 'STEP 4A - TRANSECT 1', Heading1)
    worksheet4.merge_range('D2:E2', '', Heading7)

    # Set up Transect 2 worksheet #

    worksheet5 = workbook.add_worksheet('Step 4B - Transect 2')

    worksheet5.set_default_row(56.25)

    # Add item headings to cells as strings

    x = range(1, 101)

    worksheet5.write_string('B2', 'TRANSECT 2', Heading2)
    worksheet5.write_string('C2', 'START POINT', Heading2)
    worksheet5.write_string('B3', 'GROUND LAYER', Heading4)
    worksheet5.write_string('C3', 'BELOW', Heading4)
    worksheet5.write_string('D3', 'ABOVE', Heading4)
    worksheet5.write_column('A4', x, Heading4)
    worksheet5.write('A1', None, ColorFill)
    worksheet5.write('A2', None, ColorFill)
    worksheet5.write('A3', None, ColorFill)

    # set column width

    worksheet5.set_column('A:A', 9.09)
    worksheet5.set_column('B:B', 48.00)
    worksheet5.set_column('C:C', 48.00)
    worksheet5.set_column('D:D', 48.00)
    worksheet5.set_column('E:E', 13.00)

    # Add blank cells with formatting to contain data

    worksheet5.write_column('B4', x, Heading7)
    worksheet5.write_column('C4', x, Heading7)
    worksheet5.write_column('D4', x, Heading7)

    # merge the relevant cells

    worksheet5.merge_range('B1:E1', 'STEP 4B - TRANSECT 2', Heading1)
    worksheet5.merge_range('D2:E2', '', Heading7)

    # Set up Transect 3 worksheet #

    worksheet6 = workbook.add_worksheet('Step 4C - Transect 3')

    worksheet6.set_default_row(56.25)

    # Add item headings to cells as strings

    x = range(1, 101)

    worksheet6.write_string('B2', 'TRANSECT 3', Heading2)
    worksheet6.write_string('C2', 'START POINT', Heading2)
    worksheet6.write_string('B3', 'GROUND LAYER', Heading4)
    worksheet6.write_string('C3', 'BELOW', Heading4)
    worksheet6.write_string('D3', 'ABOVE', Heading4)
    worksheet6.write_column('A4', x, Heading4)
    worksheet6.write('A1', None, ColorFill)
    worksheet6.write('A2', None, ColorFill)
    worksheet6.write('A3', None, ColorFill)

    # set column width

    worksheet6.set_column('A:A', 9.09)
    worksheet6.set_column('B:B', 48.00)
    worksheet6.set_column('C:C', 48.00)
    worksheet6.set_column('D:D', 48.00)
    worksheet6.set_column('E:E', 13.00)

    # Add blank cells with formatting to contain data

    worksheet6.write_column('B4', x, Heading7)
    worksheet6.write_column('C4', x, Heading7)
    worksheet6.write_column('D4', x, Heading7)

    # merge the relevant cells

    worksheet6.merge_range('B1:E1', 'STEP 4C - TRANSECT 3', Heading1)
    worksheet6.merge_range('D2:E2', '', Heading7)

    # Set up Basal Sweeps worksheet #

    worksheet7 = workbook.add_worksheet('Step 5 - Basal Sweeps - Table 2')

    # Add item headings to cells as strings

    worksheet7.write_string('A3', 'Location', Heading5)
    worksheet7.write_string('A4', 'Basal factor', Heading5)
    worksheet7.write_string('B5', 'Live', Heading5)
    worksheet7.write_string('C5', 'Dead', Heading5)
    worksheet7.write_string('D5', 'Live', Heading5)
    worksheet7.write_string('E5', 'Dead', Heading5)
    worksheet7.write_string('F5', 'Live', Heading5)
    worksheet7.write_string('G5', 'Dead', Heading5)
    worksheet7.write_string('H5', 'Live', Heading5)
    worksheet7.write_string('I5', 'Dead', Heading5)
    worksheet7.write_string('J5', 'Live', Heading5)
    worksheet7.write_string('K5', 'Dead', Heading5)
    worksheet7.write_string('L5', 'Live', Heading5)
    worksheet7.write_string('M5', 'Dead', Heading5)
    worksheet7.write_string('N5', 'Live', Heading5)
    worksheet7.write_string('O5', 'Dead', Heading5)
    worksheet7.write_string('A7', 'TREES', Heading4)
    worksheet7.write_string('A8', 'SHRUBS', Heading4)
    worksheet7.write_string('A17', 'ADULT TREES (>2m)', Heading4)
    worksheet7.write_string('A18', 'ADULT SHRUBS (>2m)', Heading4)
    worksheet7.write_string('A19', 'TOTAL', Heading4)
    worksheet7.write('A5', None, ColorFill)
    worksheet7.write('A6', None, ColorFill)
    worksheet7.write('A16', None, ColorFill)

    # Add blank cells with formatting to contain data

    worksheet7.write_blank('B7', None, Heading7)
    worksheet7.write_blank('C7', None, Heading7)
    worksheet7.write_blank('D7', None, Heading7)
    worksheet7.write_blank('E7', None, Heading7)
    worksheet7.write_blank('F7', None, Heading7)
    worksheet7.write_blank('G7', None, Heading7)
    worksheet7.write_blank('H7', None, Heading7)
    worksheet7.write_blank('I7', None, Heading7)
    worksheet7.write_blank('J7', None, Heading7)
    worksheet7.write_blank('K7', None, Heading7)
    worksheet7.write_blank('L7', None, Heading7)
    worksheet7.write_blank('M7', None, Heading7)
    worksheet7.write_blank('N7', None, Heading7)
    worksheet7.write_blank('O7', None, Heading7)
    worksheet7.write_blank('B8', None, Heading7)
    worksheet7.write_blank('C8', None, Heading7)
    worksheet7.write_blank('D8', None, Heading7)
    worksheet7.write_blank('E8', None, Heading7)
    worksheet7.write_blank('F8', None, Heading7)
    worksheet7.write_blank('G8', None, Heading7)
    worksheet7.write_blank('H8', None, Heading7)
    worksheet7.write_blank('I8', None, Heading7)
    worksheet7.write_blank('J8', None, Heading7)
    worksheet7.write_blank('K8', None, Heading7)
    worksheet7.write_blank('L8', None, Heading7)
    worksheet7.write_blank('M8', None, Heading7)
    worksheet7.write_blank('N8', None, Heading7)
    worksheet7.write_blank('O8', None, Heading7)

    # merge the relevant cells

    worksheet7.merge_range('A1:K1', 'STEP 5 - BASAL SWEEPS', Heading1)
    worksheet7.merge_range('L2:O2', '', Heading7)
    worksheet7.merge_range('A2:K2', 'Does site have recordable basal area?', Heading4)
    worksheet7.merge_range('B3:C3', 'North', Heading5)
    worksheet7.merge_range('D3:E3', 'Centre', Heading5)
    worksheet7.merge_range('F3:G3', 'South', Heading5)
    worksheet7.merge_range('H3:I3', 'South East', Heading5)
    worksheet7.merge_range('J3:K3', 'North West', Heading5)
    worksheet7.merge_range('L3:M3', 'North East', Heading5)
    worksheet7.merge_range('N3:O3', 'South West', Heading5)
    worksheet7.merge_range('B4:C4', '', Heading7)
    worksheet7.merge_range('D4:E4', '', Heading7)
    worksheet7.merge_range('F4:G4', '', Heading7)
    worksheet7.merge_range('H4:I4', '', Heading7)
    worksheet7.merge_range('J4:K4', '', Heading7)
    worksheet7.merge_range('L4:M4', '', Heading7)
    worksheet7.merge_range('N4:O4', '', Heading7)
    worksheet7.merge_range('B16:E16', 'BASAL AREA (m2/ha)', Heading4)
    worksheet7.merge_range('B17:E17', '', Heading7)
    worksheet7.merge_range('B18:E18', '', Heading7)
    worksheet7.merge_range('B19:E19', '', Heading7)
    worksheet7.merge_range('A20:O20', '', ColorFill)
    worksheet7.merge_range('A21:O21', 'Major woody species', Heading2)
    worksheet7.merge_range('A22:E22', 'Confirmed Species name', Heading4)
    worksheet7.merge_range('F22:K22', 'Field name', Heading4)
    worksheet7.merge_range('L22:O22', 'Functional type', Heading4)
    worksheet7.merge_range('A23:E23', '', Heading8)
    worksheet7.merge_range('A24:E24', '', Heading8)
    worksheet7.merge_range('A25:E25', '', Heading8)
    worksheet7.merge_range('A26:E26', '', Heading8)
    worksheet7.merge_range('A27:E27', '', Heading8)
    worksheet7.merge_range('A28:E28', '', Heading8)
    worksheet7.merge_range('A29:E29', '', Heading8)
    worksheet7.merge_range('A30:E30', '', Heading8)
    worksheet7.merge_range('A31:E31', '', Heading8)
    worksheet7.merge_range('A32:E32', '', Heading8)
    worksheet7.merge_range('F16:O16', '', ColorFill)
    worksheet7.merge_range('F23:K23', '', Heading7)
    worksheet7.merge_range('F24:K24', '', Heading7)
    worksheet7.merge_range('F25:K25', '', Heading7)
    worksheet7.merge_range('F26:K26', '', Heading7)
    worksheet7.merge_range('F27:K27', '', Heading7)
    worksheet7.merge_range('F28:K28', '', Heading7)
    worksheet7.merge_range('F29:K29', '', Heading7)
    worksheet7.merge_range('F30:K30', '', Heading7)
    worksheet7.merge_range('F31:K31', '', Heading7)
    worksheet7.merge_range('F32:K32', '', Heading7)
    worksheet7.merge_range('L23:O23', '', Heading7)
    worksheet7.merge_range('L24:O24', '', Heading7)
    worksheet7.merge_range('L25:O25', '', Heading7)
    worksheet7.merge_range('L26:O26', '', Heading7)
    worksheet7.merge_range('L27:O27', '', Heading7)
    worksheet7.merge_range('L28:O28', '', Heading7)
    worksheet7.merge_range('L29:O29', '', Heading7)
    worksheet7.merge_range('L30:O30', '', Heading7)
    worksheet7.merge_range('L31:O31', '', Heading7)
    worksheet7.merge_range('L32:O32', '', Heading7)

    # Write default values to cells

    worksheet7.write_string('L2', 'No', Heading7)
    worksheet7.write_string('B3', 'BLANK', Heading7)
    worksheet7.write_string('D3', 'BLANK', Heading7)
    worksheet7.write_string('F3', 'BLANK', Heading7)
    worksheet7.write_string('H3', 'BLANK', Heading7)
    worksheet7.write_string('J3', 'BLANK', Heading7)
    worksheet7.write_string('L3', 'BLANK', Heading7)
    worksheet7.write_string('N3', 'BLANK', Heading7)
    worksheet7.write_string('B4', 'BLANK', Heading7)
    worksheet7.write_string('D4', 'BLANK', Heading7)
    worksheet7.write_string('F4', 'BLANK', Heading7)
    worksheet7.write_string('H4', 'BLANK', Heading7)
    worksheet7.write_string('J4', 'BLANK', Heading7)
    worksheet7.write_string('L4', 'BLANK', Heading7)
    worksheet7.write_string('N4', 'BLANK', Heading7)

    worksheet7.write_string('A23', 'BLANK', Heading7)
    worksheet7.write_string('A24', 'BLANK', Heading7)
    worksheet7.write_string('A25', 'BLANK', Heading7)
    worksheet7.write_string('A25', 'BLANK', Heading7)
    worksheet7.write_string('A26', 'BLANK', Heading7)
    worksheet7.write_string('A27', 'BLANK', Heading7)
    worksheet7.write_string('A28', 'BLANK', Heading7)
    worksheet7.write_string('A29', 'BLANK', Heading7)
    worksheet7.write_string('A30', 'BLANK', Heading7)
    worksheet7.write_string('A31', 'BLANK', Heading7)
    worksheet7.write_string('A32', 'BLANK', Heading7)
    worksheet7.write_string('F23', 'BLANK', Heading7)
    worksheet7.write_string('F24', 'BLANK', Heading7)
    worksheet7.write_string('F25', 'BLANK', Heading7)
    worksheet7.write_string('F25', 'BLANK', Heading7)
    worksheet7.write_string('F26', 'BLANK', Heading7)
    worksheet7.write_string('F27', 'BLANK', Heading7)
    worksheet7.write_string('F28', 'BLANK', Heading7)
    worksheet7.write_string('F29', 'BLANK', Heading7)
    worksheet7.write_string('F30', 'BLANK', Heading7)
    worksheet7.write_string('F31', 'BLANK', Heading7)
    worksheet7.write_string('F32', 'BLANK', Heading7)
    worksheet7.write_string('L23', 'BLANK', Heading7)
    worksheet7.write_string('L24', 'BLANK', Heading7)
    worksheet7.write_string('L25', 'BLANK', Heading7)
    worksheet7.write_string('L25', 'BLANK', Heading7)
    worksheet7.write_string('L26', 'BLANK', Heading7)
    worksheet7.write_string('L27', 'BLANK', Heading7)
    worksheet7.write_string('L28', 'BLANK', Heading7)
    worksheet7.write_string('L29', 'BLANK', Heading7)
    worksheet7.write_string('L30', 'BLANK', Heading7)
    worksheet7.write_string('L31', 'BLANK', Heading7)
    worksheet7.write_string('L32', 'BLANK', Heading7)

    worksheet7.write_number('B7', 0, Heading7)
    worksheet7.write_number('C7', 0, Heading7)
    worksheet7.write_number('D7', 0, Heading7)
    worksheet7.write_number('E7', 0, Heading7)
    worksheet7.write_number('F7', 0, Heading7)
    worksheet7.write_number('G7', 0, Heading7)
    worksheet7.write_number('H7', 0, Heading7)
    worksheet7.write_number('I7', 0, Heading7)
    worksheet7.write_number('J7', 0, Heading7)
    worksheet7.write_number('K7', 0, Heading7)
    worksheet7.write_number('L7', 0, Heading7)
    worksheet7.write_number('M7', 0, Heading7)
    worksheet7.write_number('N7', 0, Heading7)
    worksheet7.write_number('O7', 0, Heading7)
    worksheet7.write_number('B8', 0, Heading7)
    worksheet7.write_number('C8', 0, Heading7)
    worksheet7.write_number('D8', 0, Heading7)
    worksheet7.write_number('E8', 0, Heading7)
    worksheet7.write_number('F8', 0, Heading7)
    worksheet7.write_number('G8', 0, Heading7)
    worksheet7.write_number('H8', 0, Heading7)
    worksheet7.write_number('I8', 0, Heading7)
    worksheet7.write_number('J8', 0, Heading7)
    worksheet7.write_number('K8', 0, Heading7)
    worksheet7.write_number('L8', 0, Heading7)
    worksheet7.write_number('M8', 0, Heading7)
    worksheet7.write_number('N8', 0, Heading7)
    worksheet7.write_number('O8', 0, Heading7)

    # Set column width

    worksheet7.set_column('A:A', 31.00)
    worksheet7.set_column('B:B', 13.50)
    worksheet7.set_column('C:C', 13.50)
    worksheet7.set_column('D:D', 13.50)
    worksheet7.set_column('E:E', 13.50)
    worksheet7.set_column('F:F', 13.50)
    worksheet7.set_column('G:G', 13.50)
    worksheet7.set_column('H:H', 13.50)
    worksheet7.set_column('I:I', 13.50)
    worksheet7.set_column('J:J', 13.50)
    worksheet7.set_column('K:K', 13.50)
    worksheet7.set_column('L:L', 13.50)
    worksheet7.set_column('M:M', 13.50)
    worksheet7.set_column('N:N', 13.50)
    worksheet7.set_column('O:O', 13.50)

    # set row height

    worksheet7.set_row(0, 80.25)
    worksheet7.set_row(1, 56.25)
    worksheet7.set_row(2, 33.75)
    worksheet7.set_row(3, 27.75)
    worksheet7.set_row(4, 27.75)
    worksheet7.set_row(5, 5.00)
    worksheet7.set_row(6, 56.25)
    worksheet7.set_row(7, 56.25)
    worksheet7.set_row(8, 3.00)
    worksheet7.set_row(9, 3.00)
    worksheet7.set_row(10, 3.00)
    worksheet7.set_row(11, 3.00)
    worksheet7.set_row(12, 3.00)
    worksheet7.set_row(13, 3.00)
    worksheet7.set_row(14, 3.00)
    worksheet7.set_row(15, 56.25)
    worksheet7.set_row(16, 56.25)
    worksheet7.set_row(17, 56.25)
    worksheet7.set_row(18, 56.25)
    worksheet7.set_row(19, 56.25)
    worksheet7.set_row(20, 56.25)
    worksheet7.set_row(21, 56.25)
    worksheet7.set_row(22, 56.25)
    worksheet7.set_row(23, 56.25)
    worksheet7.set_row(24, 56.25)
    worksheet7.set_row(25, 56.25)
    worksheet7.set_row(26, 56.25)
    worksheet7.set_row(27, 56.25)
    worksheet7.set_row(28, 56.25)
    worksheet7.set_row(29, 56.25)
    worksheet7.set_row(30, 56.25)
    worksheet7.set_row(31, 56.25)

    # Set up Juvenile stem count worksheet #

    worksheet8 = workbook.add_worksheet('STEP 6 - Juvenile stem count - ')

    # Add item headings to cells as strings

    worksheet8.write_string('A3', 'BELT WIDTH', Heading3)
    worksheet8.write_string('A4', 'Category', Heading5)
    worksheet8.write_string('B4', 'Transect 1', Heading5)
    worksheet8.write_string('C4', 'Transect 2', Heading5)
    worksheet8.write_string('D4', 'Transect 3', Heading5)
    worksheet8.write_string('E4', 'Transect 4', Heading5)
    worksheet8.write_string('F4', 'Transect 5', Heading5)
    worksheet8.write_string('A5', 'Juvenile shrubs (<0.5m)', Heading4)
    worksheet8.write_string('A10', 'Juvenile shrubs (<0.5m)', Heading4)
    worksheet8.write_string('B10', 'Total density (stems/ha)', Heading4)
    worksheet8.write_string('D10', 'Density class', Heading4)
    worksheet8.write_string('A6', 'Juvenile trees (<2m)', Heading4)
    worksheet8.write_string('A7', 'Total', Heading4)
    worksheet8.write_string('A11', 'Juvenile trees (<2m)', Heading4)
    worksheet8.write_string('B11', 'Total density (stem/ha)', Heading4)
    worksheet8.write_string('D11', 'Density class', Heading4)
    worksheet8.write_string('A12', 'Total', Heading4)
    worksheet8.write_string('B12', 'Total density (stems/ha)', Heading4)
    worksheet8.write_string('D12', 'Density class', Heading4)
    worksheet8.write_string('A14', "Confirmed Species Name", Heading4)

    worksheet8.write('F10', None, ColorFill)
    worksheet8.write('F11', None, ColorFill)
    worksheet8.write('F12', None, ColorFill)
    worksheet8.write('F14', None, ColorFill)
    worksheet8.write('F15', None, ColorFill)
    worksheet8.write('F16', None, ColorFill)
    worksheet8.write('F17', None, ColorFill)
    worksheet8.write('F18', None, ColorFill)
    worksheet8.write('F19', None, ColorFill)
    worksheet8.write('F20', None, ColorFill)
    worksheet8.write('F21', None, ColorFill)
    worksheet8.write('F22', None, ColorFill)
    worksheet8.write('F23', None, ColorFill)
    worksheet8.write('F24', None, ColorFill)

    # Add blank cells with formatting to contain data

    worksheet8.write_blank('A15', None, Heading7)
    worksheet8.write_blank('A16', None, Heading7)
    worksheet8.write_blank('A17', None, Heading7)
    worksheet8.write_blank('A18', None, Heading7)
    worksheet8.write_blank('A19', None, Heading7)
    worksheet8.write_blank('A20', None, Heading7)
    worksheet8.write_blank('A21', None, Heading7)
    worksheet8.write_blank('A22', None, Heading7)
    worksheet8.write_blank('A23', None, Heading7)
    worksheet8.write_blank('A24', None, Heading7)

    # Merge the relevant cells

    worksheet8.merge_range('A1:F1', 'STEP 6 - JUVENILE STEM COUNT', Heading1)
    worksheet8.merge_range('A2:E2', 'Does site show appreciable thickening?', Heading2)
    worksheet8.merge_range('B3:F3', '', Heading7)
    worksheet8.merge_range('A13:F13', 'MAJOR WOODY SPECIES', Heading2)
    worksheet8.merge_range('B14:C14', 'Field Name', Heading4)
    worksheet8.merge_range('B15:C15', '', Heading7)
    worksheet8.merge_range('D14:E14', 'Functional Type', Heading4)
    worksheet8.merge_range('D15:E15', '', Heading7)
    worksheet8.merge_range('B16:C16', '', Heading7)
    worksheet8.merge_range('D16:E16', '', Heading7)
    worksheet8.merge_range('B17:C17', '', Heading7)
    worksheet8.merge_range('D17:E17', '', Heading7)
    worksheet8.merge_range('B18:C18', '', Heading7)
    worksheet8.merge_range('D18:E18', '', Heading7)
    worksheet8.merge_range('B19:C19', '', Heading7)
    worksheet8.merge_range('D19:E19', '', Heading7)
    worksheet8.merge_range('B20:C20', '', Heading7)
    worksheet8.merge_range('D20:E20', '', Heading7)
    worksheet8.merge_range('B21:C21', '', Heading7)
    worksheet8.merge_range('D21:E21', '', Heading7)
    worksheet8.merge_range('B22:C22', '', Heading7)
    worksheet8.merge_range('D22:E22', '', Heading7)
    worksheet8.merge_range('B23:C23', '', Heading7)
    worksheet8.merge_range('D23:E23', '', Heading7)
    worksheet8.merge_range('B24:C24', '', Heading7)
    worksheet8.merge_range('D24:E24', '', Heading7)

    # Write default values to cells

    worksheet8.write_number('C10', 0, Heading7)
    worksheet8.write_number('C11', 0, Heading7)
    worksheet8.write_number('C12', 0, Heading7)
    worksheet8.write_string('E10', 'Not observed', Heading7)
    worksheet8.write_string('E11', 'Not observed', Heading7)
    worksheet8.write_string('E12', 'Not observed', Heading7)
    worksheet8.write_string('F2', 'No', Heading7)
    worksheet8.write_string('B3', 'BLANK', Heading7)
    worksheet8.write_number('B5', 0, Heading7)
    worksheet8.write_number('C5', 0, Heading7)
    worksheet8.write_number('D5', 0, Heading7)
    worksheet8.write_number('E5', 0, Heading7)
    worksheet8.write_number('F5', 0, Heading7)
    worksheet8.write_number('B6', 0, Heading7)
    worksheet8.write_number('C6', 0, Heading7)
    worksheet8.write_number('D6', 0, Heading7)
    worksheet8.write_number('E6', 0, Heading7)
    worksheet8.write_number('F6', 0, Heading7)
    worksheet8.write_number('B7', 0, Heading7)
    worksheet8.write_number('C7', 0, Heading7)
    worksheet8.write_number('D7', 0, Heading7)
    worksheet8.write_number('E7', 0, Heading7)
    worksheet8.write_number('F7', 0, Heading7)

    worksheet8.write_string('D15', 'BLANK', Heading7)
    worksheet8.write_string('D16', 'BLANK', Heading7)
    worksheet8.write_string('D17', 'BLANK', Heading7)
    worksheet8.write_string('D18', 'BLANK', Heading7)
    worksheet8.write_string('D19', 'BLANK', Heading7)
    worksheet8.write_string('D20', 'BLANK', Heading7)
    worksheet8.write_string('D21', 'BLANK', Heading7)
    worksheet8.write_string('D22', 'BLANK', Heading7)
    worksheet8.write_string('D23', 'BLANK', Heading7)
    worksheet8.write_string('D24', 'BLANK', Heading7)

    # Set column width

    worksheet8.set_column('A:A', 40.00)
    worksheet8.set_column('B:B', 15.00)
    worksheet8.set_column('C:C', 15.00)
    worksheet8.set_column('D:D', 15.00)
    worksheet8.set_column('E:E', 15.00)
    worksheet8.set_column('F:F', 15.00)

    # set row height

    worksheet8.set_row(0, 47.25)
    worksheet8.set_row(1, 47.25)
    worksheet8.set_row(2, 47.25)
    worksheet8.set_row(3, 33.75)
    worksheet8.set_row(4, 56.25)
    worksheet8.set_row(5, 56.25)
    worksheet8.set_row(6, 56.25)
    worksheet8.set_row(7, 3.00)
    worksheet8.set_row(8, 3.00)
    worksheet8.set_row(9, 65.00)
    worksheet8.set_row(10, 65.00)
    worksheet8.set_row(11, 65.00)
    worksheet8.set_row(12, 65.00)
    worksheet8.set_row(13, 65.00)
    worksheet8.set_row(14, 65.00)
    worksheet8.set_row(15, 65.00)
    worksheet8.set_row(16, 65.00)
    worksheet8.set_row(17, 65.00)
    worksheet8.set_row(18, 65.00)
    worksheet8.set_row(19, 65.00)
    worksheet8.set_row(20, 65.00)
    worksheet8.set_row(21, 65.00)
    worksheet8.set_row(22, 65.00)
    worksheet8.set_row(23, 65.00)

    # Set up Ground layer composition worksheet #

    worksheet9 = workbook.add_worksheet('Output 1 - Ground layer composi')

    # Set row height

    worksheet9.set_default_row(56.25)

    # Add item headings to cells as strings

    worksheet9.write_string('B5', 'Total site %', Heading4)
    worksheet9.write_string('C5', 'Vegetation cover %', Heading4)
    worksheet9.write_string('D5', 'Total site %', Heading4)
    worksheet9.write_string('E5', 'Vegetation cover %', Heading4)
    worksheet9.write_string('A6', 'Perennial grass', Heading4)
    worksheet9.write_string('A7', 'Annual grass', Heading4)
    worksheet9.write_string('A8', 'Perennial forb', Heading4)
    worksheet9.write_string('A9', 'Annual forb', Heading4)
    worksheet9.write_string('A10', 'Unspecified plant', Heading4)
    worksheet9.write_string('A11', 'Total veg', Heading4)
    worksheet9.write_string('A12', 'Litter', Heading4)
    worksheet9.write_string('A13', 'Bare Ground', Heading4)
    worksheet9.write('A4', None, ColorFill)
    worksheet9.write('A5', None, ColorFill)
    worksheet9.write('A2', None, ColorFill)

    # Set column width

    worksheet9.set_column('A:A', 30.27)
    worksheet9.set_column('B:B', 30.27)
    worksheet9.set_column('C:C', 30.27)
    worksheet9.set_column('D:D', 30.27)
    worksheet9.set_column('E:E', 30.27)

    # Add blank cells with formatting to contain data

    worksheet9.write_blank('C3', None, Heading7)
    worksheet9.write_blank('B6', None, Heading7)
    worksheet9.write_blank('B7', None, Heading7)
    worksheet9.write_blank('B8', None, Heading7)
    worksheet9.write_blank('B9', None, Heading7)
    worksheet9.write_blank('B10', None, Heading7)
    worksheet9.write_blank('B11', None, Heading7)
    worksheet9.write_blank('B12', None, Heading7)
    worksheet9.write_blank('B13', None, Heading7)
    worksheet9.write_blank('C6', None, Heading7)
    worksheet9.write_blank('C7', None, Heading7)
    worksheet9.write_blank('C8', None, Heading7)
    worksheet9.write_blank('C9', None, Heading7)
    worksheet9.write_blank('C10', None, Heading7)
    worksheet9.write_blank('D6', None, Heading7)
    worksheet9.write_blank('D7', None, Heading7)
    worksheet9.write_blank('D8', None, Heading7)
    worksheet9.write_blank('D9', None, Heading7)
    worksheet9.write_blank('D10', None, Heading7)
    worksheet9.write_blank('D11', None, Heading7)
    worksheet9.write_blank('D12', None, Heading7)
    worksheet9.write_blank('D13', None, Heading7)
    worksheet9.write_blank('E6', None, Heading7)
    worksheet9.write_blank('E7', None, Heading7)
    worksheet9.write_blank('E8', None, Heading7)
    worksheet9.write_blank('E9', None, Heading7)
    worksheet9.write_blank('E10', None, Heading7)

    # merge the relevant cells

    worksheet9.merge_range('A1:E1', 'GROUND LAYER COMPOSITION', Heading1)
    worksheet9.merge_range('A3:B3', 'DO THE TRANSECTS ACCURATELY REPRESENT THE SITE?', Heading4)
    worksheet9.merge_range('B4:C4', 'Transect data', Heading4)
    worksheet9.merge_range('D4:E4', 'Cover estimates', Heading4)

    # Set up Cover estimates worksheet #

    worksheet10 = workbook.add_worksheet('Step 7 - Cover estimates  - Tab')

    # Set row height

    worksheet10.set_default_row(56.25)

    # Add item headings to cells as strings

    worksheet10.write_string('B3', 'LITTER', Heading2)
    worksheet10.write_string('C3', 'BARE GROUND', Heading2)
    worksheet10.write_string('D3', 'TOTAL VEG', Heading2)
    worksheet10.write_string('A4', 'Transect figure', Heading2)
    worksheet10.write_string('A5', 'Adjusted est', Heading2)
    worksheet10.write_string('B7', 'Perennial grasses', Heading4)
    worksheet10.write_string('C7', 'Annual grasses', Heading4)
    worksheet10.write_string('D7', 'Perennial forbs', Heading4)
    worksheet10.write_string('E7', 'Annual forbs', Heading4)
    worksheet10.write_string('F7', 'Unspecified plants', Heading4)
    worksheet10.write_string('A8', 'Transect figure', Heading3)
    worksheet10.write_string('A9', 'Adjusted figure', Heading3)
    worksheet10.write_string('A10', 'Estimate total', Heading3)
    worksheet10.write_string('C14', 'Confirmed species name', Heading4)
    worksheet10.write_string('D14', 'Field name', Heading4)
    worksheet10.write_string('E14', 'Cover estimate', Heading4)
    worksheet10.write_string('C25', 'Other species', Heading6)
    worksheet10.write_string('D25', 'Other species', Heading6)
    worksheet10.write_string('C27', 'Confirmed species name', Heading4)
    worksheet10.write_string('D27', 'Field name', Heading4)
    worksheet10.write_string('E27', 'Cover estimate', Heading4)
    worksheet10.write_string('C32', 'Other species', Heading6)
    worksheet10.write_string('D32', 'Other species', Heading6)
    worksheet10.write_string('C34', 'Confirmed species name', Heading4)
    worksheet10.write_string('D34', 'Field name', Heading4)
    worksheet10.write_string('E34', 'Cover estimate', Heading4)
    worksheet10.write_string('C39', 'Other species', Heading6)
    worksheet10.write_string('D39', 'Other species', Heading6)
    worksheet10.write_string('C41', 'Confirmed species name', Heading4)
    worksheet10.write_string('D41', 'Field name', Heading4)
    worksheet10.write_string('E41', 'Cover estimate', Heading4)
    worksheet10.write_string('C46', 'Other species', Heading6)
    worksheet10.write_string('D46', 'Other species', Heading6)
    worksheet10.write_string('C48', 'Confirmed species name', Heading4)
    worksheet10.write_string('D48', 'Field name', Heading4)
    worksheet10.write_string('E48', 'Cover estimate', Heading4)
    worksheet10.write_string('C53', 'Other species', Heading6)
    worksheet10.write_string('D53', 'Other species', Heading6)

    worksheet10.write('A3', None, ColorFill)
    worksheet10.write('A7', None, ColorFill)

    # merge the relevant cells

    worksheet10.merge_range('A1:F1', 'STEP 7 - COVER ESTIMATES', Heading1)
    worksheet10.merge_range('A2:F2', 'SITE COVER FRACTIONS', Heading1)
    worksheet10.merge_range('E3:F3', 'SUM', Heading2)
    worksheet10.merge_range('E4:F4', '', Heading7)
    worksheet10.merge_range('E5:F5', '', Heading7)
    worksheet10.merge_range('A6:F6', 'VEGETATION COVER', Heading1)
    worksheet10.merge_range('B10:F10', '', Heading7)
    worksheet10.merge_range('A11:F11', '')
    worksheet10.merge_range('A12:F12', 'VEGETATION', Heading1)
    worksheet10.merge_range('A13:E13', 'Perennial grasses', Heading1)
    worksheet10.merge_range('A14:B14', 'PROPORTION OF VEG COVER', Heading4)
    worksheet10.merge_range('C15:E15', '3P grasses', Heading1)
    worksheet10.merge_range('C20:E20', 'Other perennial grasses', Heading1)
    worksheet10.merge_range('A15:B24', '', Heading7)
    worksheet10.merge_range('A25:B25', '', ColorFill)
    worksheet10.merge_range('A26:E26', 'Annual grasses', Heading1)
    worksheet10.merge_range('A27:B27', 'PROPORTION OF VEG COVER', Heading4)
    worksheet10.merge_range('A28:B31', '', Heading7)
    worksheet10.merge_range('A32:B32', '', ColorFill)
    worksheet10.merge_range('A33:E33', 'Perennial forbs', Heading1)
    worksheet10.merge_range('A34:B34', 'PROPORTION OF VEG COVER', Heading4)
    worksheet10.merge_range('A35:B38', '', Heading7)
    worksheet10.merge_range('A39:B39', '', ColorFill)
    worksheet10.merge_range('A40:E40', 'Annual forbs', Heading1)
    worksheet10.merge_range('A41:B41', 'PROPORTION OF VEG COVER', Heading4)
    worksheet10.merge_range('A42:B45', '', Heading7)
    worksheet10.merge_range('A46:B46', '', ColorFill)
    worksheet10.merge_range('A47:E47', 'Unspecified plants', Heading1)
    worksheet10.merge_range('A48:B48', 'PROPORTION OF VEG COVER', Heading4)
    worksheet10.merge_range('A49:B52', '', Heading7)
    worksheet10.merge_range('A53:B53', '', ColorFill)

    # Add blank cells with formatting to contain data

    worksheet10.write_blank('B4', None, Heading7)
    worksheet10.write_blank('C4', None, Heading7)
    worksheet10.write_blank('D4', None, Heading7)
    worksheet10.write_blank('B8', None, Heading7)
    worksheet10.write_blank('C8', None, Heading7)
    worksheet10.write_blank('D8', None, Heading7)
    worksheet10.write_blank('E8', None, Heading7)
    worksheet10.write_blank('F8', None, Heading7)
    worksheet10.write_blank('C16', None, Heading7)
    worksheet10.write_blank('D16', None, Heading7)
    worksheet10.write_blank('E16', None, Heading7)
    worksheet10.write_blank('C17', None, Heading7)
    worksheet10.write_blank('D17', None, Heading7)
    worksheet10.write_blank('E17', None, Heading7)
    worksheet10.write_blank('C18', None, Heading7)
    worksheet10.write_blank('D18', None, Heading7)
    worksheet10.write_blank('E18', None, Heading7)
    worksheet10.write_blank('C19', None, Heading7)
    worksheet10.write_blank('D19', None, Heading7)
    worksheet10.write_blank('E19', None, Heading7)
    worksheet10.write_blank('C21', None, Heading7)
    worksheet10.write_blank('D21', None, Heading7)
    worksheet10.write_blank('E21', None, Heading7)
    worksheet10.write_blank('C22', None, Heading7)
    worksheet10.write_blank('D22', None, Heading7)
    worksheet10.write_blank('E22', None, Heading7)
    worksheet10.write_blank('C23', None, Heading7)
    worksheet10.write_blank('D23', None, Heading7)
    worksheet10.write_blank('E23', None, Heading7)
    worksheet10.write_blank('C24', None, Heading7)
    worksheet10.write_blank('D24', None, Heading7)
    worksheet10.write_blank('E24', None, Heading7)
    worksheet10.write_blank('C28', None, Heading7)
    worksheet10.write_blank('D28', None, Heading7)
    worksheet10.write_blank('E28', None, Heading7)
    worksheet10.write_blank('C29', None, Heading7)
    worksheet10.write_blank('D29', None, Heading7)
    worksheet10.write_blank('E29', None, Heading7)
    worksheet10.write_blank('C30', None, Heading7)
    worksheet10.write_blank('D30', None, Heading7)
    worksheet10.write_blank('E30', None, Heading7)
    worksheet10.write_blank('C31', None, Heading7)
    worksheet10.write_blank('D31', None, Heading7)
    worksheet10.write_blank('E31', None, Heading7)
    worksheet10.write_blank('C35', None, Heading7)
    worksheet10.write_blank('D35', None, Heading7)
    worksheet10.write_blank('E35', None, Heading7)
    worksheet10.write_blank('C36', None, Heading7)
    worksheet10.write_blank('D36', None, Heading7)
    worksheet10.write_blank('E36', None, Heading7)
    worksheet10.write_blank('C37', None, Heading7)
    worksheet10.write_blank('D37', None, Heading7)
    worksheet10.write_blank('E37', None, Heading7)
    worksheet10.write_blank('C38', None, Heading7)
    worksheet10.write_blank('D38', None, Heading7)
    worksheet10.write_blank('E38', None, Heading7)
    worksheet10.write_blank('C42', None, Heading7)
    worksheet10.write_blank('D42', None, Heading7)
    worksheet10.write_blank('E42', None, Heading7)
    worksheet10.write_blank('C43', None, Heading7)
    worksheet10.write_blank('D43', None, Heading7)
    worksheet10.write_blank('E43', None, Heading7)
    worksheet10.write_blank('C44', None, Heading7)
    worksheet10.write_blank('D44', None, Heading7)
    worksheet10.write_blank('E44', None, Heading7)
    worksheet10.write_blank('C45', None, Heading7)
    worksheet10.write_blank('D45', None, Heading7)
    worksheet10.write_blank('E45', None, Heading7)
    worksheet10.write_blank('C49', None, Heading7)
    worksheet10.write_blank('D49', None, Heading7)
    worksheet10.write_blank('E49', None, Heading7)
    worksheet10.write_blank('C50', None, Heading7)
    worksheet10.write_blank('D50', None, Heading7)
    worksheet10.write_blank('E50', None, Heading7)
    worksheet10.write_blank('C51', None, Heading7)
    worksheet10.write_blank('D51', None, Heading7)
    worksheet10.write_blank('E51', None, Heading7)
    worksheet10.write_blank('C52', None, Heading7)
    worksheet10.write_blank('D52', None, Heading7)
    worksheet10.write_blank('E52', None, Heading7)
    worksheet10.write_blank('E25', None, Heading7)
    worksheet10.write_blank('E32', None, Heading7)
    worksheet10.write_blank('E39', None, Heading7)
    worksheet10.write_blank('E46', None, Heading7)
    worksheet10.write_blank('E53', None, Heading7)

    # Write default values to cells

    worksheet10.write_number('B5', 0, Heading7)
    worksheet10.write_number('C5', 0, Heading7)
    worksheet10.write_number('D5', 0, Heading7)
    worksheet10.write_number('B9', 0, Heading7)
    worksheet10.write_number('C9', 0, Heading7)
    worksheet10.write_number('D9', 0, Heading7)
    worksheet10.write_number('E9', 0, Heading7)
    worksheet10.write_number('F9', 0, Heading7)
    worksheet10.write_number('C11', 0, Heading7)
    worksheet10.write_number('E16', 0, Heading7)
    worksheet10.write_number('E17', 0, Heading7)
    worksheet10.write_number('E18', 0, Heading7)
    worksheet10.write_number('E19', 0, Heading7)
    worksheet10.write_number('E21', 0, Heading7)
    worksheet10.write_number('E22', 0, Heading7)
    worksheet10.write_number('E23', 0, Heading7)
    worksheet10.write_number('E24', 0, Heading7)
    worksheet10.write_number('E28', 0, Heading7)
    worksheet10.write_number('E29', 0, Heading7)
    worksheet10.write_number('E30', 0, Heading7)
    worksheet10.write_number('E31', 0, Heading7)
    worksheet10.write_number('E35', 0, Heading7)
    worksheet10.write_number('E36', 0, Heading7)
    worksheet10.write_number('E37', 0, Heading7)
    worksheet10.write_number('E38', 0, Heading7)
    worksheet10.write_number('E42', 0, Heading7)
    worksheet10.write_number('E43', 0, Heading7)
    worksheet10.write_number('E44', 0, Heading7)
    worksheet10.write_number('E45', 0, Heading7)
    worksheet10.write_number('E49', 0, Heading7)
    worksheet10.write_number('E50', 0, Heading7)
    worksheet10.write_number('E51', 0, Heading7)
    worksheet10.write_number('E52', 0, Heading7)

    # Set column width

    worksheet10.set_column('A:A', 31.27)
    worksheet10.set_column('B:B', 34.91)
    worksheet10.set_column('C:C', 34.91)
    worksheet10.set_column('D:D', 30.55)
    worksheet10.set_column('E:E', 25.82)
    worksheet10.set_column('F:F', 24.55)

    # Set up Site Condition worksheet #

    worksheet11 = workbook.add_worksheet('Step 8 - Site Condition')

    # set row height

    worksheet11.set_default_row(56.25)

    # Add item headings to cells as strings

    worksheet11.write_string('A2', 'ITEM', Heading2)
    worksheet11.write_string('A3', 'GREENESS:', Heading3)
    worksheet11.write_string('A4', "Comments", Heading3)
    worksheet11.write_string('A5', 'ABUNDANCE:', Heading3)
    worksheet11.write_string('A6', 'Comments', Heading3)
    worksheet11.write_string('A7', 'UTILISATION:', Heading3)
    worksheet11.write_string('A8', 'Comments', Heading3)
    worksheet11.write_string('A9', 'LAND CONDITION SCORE:', Heading3)
    worksheet11.write('A10', None, ColorFill)

    # merge the relevant cells

    worksheet11.merge_range('A1:L1', 'STEP 8 - SITE CONDITION CHARACTERISTICS', Heading1)
    worksheet11.merge_range('A11:A12', 'VISIT ASSESSMENT NOTES:', Heading3)
    worksheet11.merge_range('B2:L2', 'INPUT', Heading2)
    worksheet11.merge_range('B3:L3', '', Heading7)
    worksheet11.merge_range('B4:L4', '', Heading7)
    worksheet11.merge_range('B5:L5', '', Heading7)
    worksheet11.merge_range('B6:L6', '', Heading7)
    worksheet11.merge_range('B7:L7', '', Heading7)
    worksheet11.merge_range('B8:L8', '', Heading7)
    worksheet11.merge_range('G9:L9', 'ASSESSMENT SCORE (good,fair, poor)', Heading4)
    worksheet11.merge_range('B9:F9', 'LAND COND GUIDE (A, B, C, D)', Heading4)
    worksheet11.merge_range('B10:F10', '', Heading7)
    worksheet11.merge_range('G10:L10', '', Heading7)
    worksheet11.merge_range('B11:L12', '', Heading7)
    worksheet11.merge_range('A13:L13', '', ColorFill)
    worksheet11.merge_range('B14:L14', '', Heading7)
    worksheet11.merge_range('B15:L15', '', Heading7)
    worksheet11.merge_range('B16:L16', '', Heading7)
    worksheet11.merge_range('B17:L17', '', Heading7)
    worksheet11.merge_range('B18:L18', '', Heading7)

    # Set column width

    worksheet11.set_column('A:A', 37.82)
    worksheet11.set_column('B:B', 9.09)
    worksheet11.set_column('C:C', 8.45)
    worksheet11.set_column('D:D', 10.91)
    worksheet11.set_column('E:E', 6.27)
    worksheet11.set_column('F:F', 8.00)
    worksheet11.set_column('G:G', 8.36)
    worksheet11.set_column('H:H', 10.18)
    worksheet11.set_column('I:I', 8.18)
    worksheet11.set_column('J:J', 0.61)
    worksheet11.set_column('K:K', 0.61)
    worksheet11.set_column('L:L', 0.61)

    workbook.close()


def writeTransect(path, site, final_cover_df):
    """Writes the trasnect intercept values to each site spreadsheet"""

    site_slice = final_cover_df['SITE'] == site
    site_final = final_cover_df[site_slice]

    # Extract records for each transect from dataframe

    t1_values = site_final.iloc[0:100, 3:6]
    t2_values = site_final.iloc[100:200, 3:6]
    t3_values = site_final.iloc[200:300, 3:6]

    # Set up writer for Pandas-Excel conversion

    book = load_workbook(path + site + '.xlsx')
    writer = pd.ExcelWriter(path + site + '.xlsx', engine='openpyxl')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

    # Write dataframes for the three transects to the Excel workbook

    t1_values.to_excel(writer, startrow=3, startcol=1, sheet_name="Step 4A - Transect 1", index=False, header=False)
    t2_values.to_excel(writer, startrow=3, startcol=1, sheet_name="Step 4B - Transect 2", index=False, header=False)
    t3_values.to_excel(writer, startrow=3, startcol=1, sheet_name="Step 4C - Transect 3", index=False, header=False)

    ### Write transect name to excel ###

    # Declare variables for each worksheet

    t1_sheet = book['Step 4A - Transect 1']
    t2_sheet = book['Step 4B - Transect 2']
    t3_sheet = book['Step 4C - Transect 3']

    # Create variable for the transect names:

    t1_name = site_final.iat[0, 1]
    t2_name = site_final.iat[100, 1]
    t3_name = site_final.iat[200, 1]

    # Create variables for sheet locations and write values

    t1 = t1_sheet['D2']
    t1.value = t1_name

    t2 = t2_sheet['D2']
    t2.value = t2_name

    t3 = t3_sheet['D2']
    t3.value = t3_name

    writer.save()


def createMainData(main_data, gps_photo):
    main_interim = pd.merge(main_data, gps_photo, on='SITE')

    main_final = main_interim.fillna(value='BLANK')

    # Change values in columns to required values for workbook

    landscape_values = {'sloping': 'Sloping', 'flat': 'Flat', 'in_a_depression': 'In a depression',
                        'on_a_crest': 'On a crest'}

    soil_color_values = {'black_soil': 'Black', 'grey_soil': 'Grey', 'brown_soil': 'Brown', 'red_soil': 'Red',
                         'yellow_soil': 'Yellow', 'pale_soil': 'Pale'}

    """soil_moist_values = {'wet_soil': 'Wet', 'moist_soil': 'Moist', 'dry_soil': 'Dry'}"""

    """surf_crack_values = {'absent_crack': 'Absent', 'fine_crack': 'Fine (1): less than 5 mm.',
                         'medium_crack': 'Medium: 5 to 10 mm.',
                         'coarse_crack': 'Coarse: 10 to 20 mm.', 'very_coarse_crack': 'Very coarse (4): 20 - 50 mm.',
                         'extermely_coarse_crack': 'Extremely coarse (5)'}"""

    """season_cond_values = {'recent_rainfall': 'Recent rainfall',
                          'several_months_since_rain': 'Several months since rain',
                          'extended_dry_period': 'Extended dry period'}"""

    """atm_cond_values = {'clear_sky': 'Clear sky', 'some_cloud': 'Some cloud 1- 5 % of sky',
                       'patchy_cloud': 'Patchy cloud 5 - 10% of sky',
                       'moderate_cloud': 'Moderate cloud 10 - 30% of sky',
                       'heavy_cloud': 'Heavy cloud 40 - 70% of the sky', 'dense_cloud': 'Dense cloud 70 - 100%',
                       'slight_haze': 'Slight haze of smoke or dust', 'moderate_haze': 'Moderate haze of smoke or dust',
                       'dense_haze': 'Dense haze of smoke or dust'}"""

    ls_source_values = {'north_summary_250k': 'North summary of 250k',
                        'north_interim_descriptions': 'North Interim LS descriptions',
                        'south_1_million_csiro': 'South 1 million CSIRO', 'south_250k_dlrm': 'South 250k DLRM'}

    ls_assessed_values = {'ls_yes': 'Yes', 'ls_no': 'No', 'ls_not_assess': 'Not Assessed'}

    water_type_values = {'bore': 'Bore', 'dam': 'Dam', 'trough': 'Trough', 'creek_river': 'Creek/river',
                         'waterhole': 'Waterhole', 'turkey_nest': 'Turkey nest', 'water_other': 'Other'}

    """direction_values = {'north': 'N', 'north_east': 'NE', 'east': 'E', 'south_east': 'SE',
                        'south': 'S', 'south_west': 'SW', 'west': 'W', 'north_west': 'NW'}"""

    adj_infra_values = {'road_or_track': 'Roads or tracks', 'old_track': 'Old track - not active',
                        'watering_points': 'Watering points - dams, bores, tanks', 'fence': 'Fence',
                        'adj_infra_other': 'Other'}

    true_false = {1: 'true', 0: 'false'}

    present_absent = {'present': 'Present', 'absent': 'Absent'}

    yesno_to_presabsence = {'yes': 'Present', 'no': 'Absent'}

    north_ff_values = {'NFF_absent': 'Absent', 'since_last_growth_event': 'Since last growth event',
                       'before_last_growth_event': 'Before last growth event'}

    north_fi_values = {'NFI_absent': 'Absent', 'low_intensity_cool_fire': 'Low intensity/cool fire',
                       'low_moderate': 'Low/moderate', 'moderate': 'Moderate',
                       'moderate_high': 'Moderate/high', 'high': 'High'}

    south_ff_values = {'SFF_absent': 'Absent', '<1': '<12 months', '1_2': '1-2 years', '2_10': '2-10 years',
                       '>10': '>10 years'}

    south_fi_values = {'SFI_absent': 'Absent', 'cool': 'Cool fire', 'hot': 'Hot fire'}

    erosion_severity_values = {'sev_absent': 'Absent', 'natural': 'Natural', 'slight': 'Slight', 'moderate': 'Moderate',
                               'severe': 'Severe'}

    erosion_stability_values = {'active': 'Active', 'stable': 'Stable (historical)'}

    yes_no = {'yes': 'Yes', 'no': 'No'}

    greeness_values = {'green': 'GREEN', 'drying_off_greening_up': 'DRYING OFF/GREENING UP', 'dry': 'DRY'}

    abundance_values = {'abundant': 'Abundant', 'moderately_abundant': 'Moderately abundant', 'sparse': 'Sparse'}

    propor_values = {'no_grazing': 'NO GRAZING', '<10%': '<10%', '11_25%': '11%-25%', '26_50%': '26%-50%',
                     '51_75%': '51%-75%', '76_90%': '76%-90%', '>90%': '>90%'}

    cond_class_values = {'good': 'GOOD', 'fair': 'FAIR', 'poor': 'POOR'}

    main_final['LANDSCAPE_POS'] = main_final['LANDSCAPE_POS'].replace(landscape_values)
    main_final['SOIL_COLOR'] = main_final['SOIL_COLOR'].replace(soil_color_values)
    main_final['SOIL_MOIST'] = main_final['SOIL_MOIST'].replace(soil_moist_values)
    main_final['SURF_CRACK'] = main_final['SURF_CRACK'].replace(surf_crack_values)
    main_final['SEASON_COND'] = main_final['SEASON_COND'].replace(season_cond_values)
    main_final['ATM_COND'] = main_final['ATM_COND'].replace(atm_cond_values)
    main_final['LAND_SYS_SOURCE'] = main_final['LAND_SYS_SOURCE'].replace(ls_source_values)
    main_final['LAND_SYS_CONSIST'] = main_final['LAND_SYS_CONSIST'].replace(ls_assessed_values)
    main_final['WATER_TYPE'] = main_final['WATER_TYPE'].replace(water_type_values)
    main_final['DIRECTION_WATER'] = main_final['DIRECTION_WATER'].replace(direction_values)
    main_final['DIRECTION_TRACK'] = main_final['DIRECTION_TRACK'].replace(direction_values)
    main_final['ADJACENT_INFRASTRUCTURE'] = main_final['ADJACENT_INFRASTRUCTURE'].replace(adj_infra_values)
    main_final['CAMEL'] = main_final['CAMEL'].replace(true_false)
    main_final['RABBIT'] = main_final['RABBIT'].replace(true_false)
    main_final['DONKEY'] = main_final['DONKEY'].replace(true_false)
    main_final['HORSE'] = main_final['HORSE'].replace(true_false)
    main_final['PIG'] = main_final['PIG'].replace(true_false)
    main_final['BUFFALO'] = main_final['BUFFALO'].replace(true_false)
    main_final['NATIVE_HERBIVORE'] = main_final['NATIVE_HERBIVORE'].replace(true_false)
    main_final['OTHER'] = main_final['OTHER'].replace(true_false)
    main_final['CATTLE_PAD'] = main_final['CATTLE_PAD'].replace(present_absent)
    main_final['CATTLE_TRAMP'] = main_final['CATTLE_TRAMP'].replace(present_absent)
    main_final['NORTH_FF'] = main_final['NORTH_FF'].replace(north_ff_values)
    main_final['NORTH_FI'] = main_final['NORTH_FI'].replace(north_fi_values)
    main_final['SOUTH_FF'] = main_final['SOUTH_FF'].replace(south_ff_values)
    main_final['SOUTH_FI'] = main_final['SOUTH_FI'].replace(south_fi_values)
    main_final['ERODIBLE_SOIL'] = main_final['ERODIBLE_SOIL'].replace(yes_no)
    main_final['SCALDING_SEVERITY'] = main_final['SCALDING_SEVERITY'].replace(erosion_severity_values)
    main_final['SCALDING_STABILITY'] = main_final['SCALDING_STABILITY'].replace(erosion_stability_values)
    main_final['WINDSHEETING_SEVERITY'] = main_final['WINDSHEETING_SEVERITY'].replace(erosion_severity_values)
    main_final['WINDSHEETING_STABILITY'] = main_final['WINDSHEETING_STABILITY'].replace(erosion_stability_values)
    main_final['WATERSHEETING_SEVERITY'] = main_final['WATERSHEETING_SEVERITY'].replace(erosion_severity_values)
    main_final['WATERSHEETING_STABILITY'] = main_final['WATERSHEETING_STABILITY'].replace(erosion_stability_values)
    main_final['RILLING_SEVERITY'] = main_final['RILLING_SEVERITY'].replace(erosion_severity_values)
    main_final['RILLING_STABILITY'] = main_final['RILLING_STABILITY'].replace(erosion_stability_values)
    main_final['GULLYING_SEVERITY'] = main_final['GULLYING_SEVERITY'].replace(erosion_severity_values)
    main_final['GULLYING_STABILITY'] = main_final['GULLYING_STABILITY'].replace(erosion_stability_values)
    main_final['CLEARING'] = main_final['CLEARING'].replace(yesno_to_presabsence)
    main_final['CYCLONE_STORM'] = main_final['CYCLONE_STORM'].replace(yesno_to_presabsence)
    main_final['DIEBACK'] = main_final['DIEBACK'].replace(yesno_to_presabsence)
    main_final['ERODIABLE_SOIL'] = main_final['ERODIBLE_SOIL'].replace(yes_no)
    main_final['GREENESS'] = main_final['GREENESS'].replace(greeness_values)
    main_final['ABUNDANCE'] = main_final['ABUNDANCE'].replace(abundance_values)
    main_final['PAST_UTIL_PROP'] = main_final['PAST_UTIL_PROP'].replace(propor_values)
    main_final['CONDITION_CLASS'] = main_final['CONDITION_CLASS'].replace(cond_class_values)

    return main_final


def createSiteMainData(site, main_final):
    df1 = main_final['SITE'] == site
    main_site_final = main_final[df1]

    return main_site_final


def createTransectInfo(transect_basic):
    # Set blank adjusted cover fields to 0
    transect_basic[
        ['LITTER_ADJUSTED', 'EXPOSED_GROUND_ADJUSTED', 'TOTAL_VEG_ADJUSTED', 'PG_TOTAL_ADJUSTED', 'AG_TOTAL_ADJUSTED',
         'PF_TOTAL_ADJUSTED', 'AF_TOTAL_ADJUSTED', 'UP_TOTAL_ADJUSTED', 'PG_3P_SP1_COVER', 'PG_3P_SP2_COVER',
         'PG_3P_SP3_COVER', 'PG_3P_SP4_COVER', 'PG_OTHER_SP1_COVER', 'PG_OTHER_SP2_COVER', 'PG_OTHER_SP3_COVER',
         'PG_OTHER_SP4_COVER', 'AG_SP1_COVER', 'AG_SP2_COVER', 'AG_SP3_COVER', 'AG_SP4_COVER', 'PF_SP1_COVER',
         'PF_SP2_COVER', 'PF_SP3_COVER', 'PF_SP4_COVER', 'AF_SP1_COVER', 'AF_SP2_COVER', 'AF_SP3_COVER',
         'AF_SP4_COVER', 'UP_SP1_COVER', 'UP_SP2_COVER', 'UP_SP3_COVER', 'UP_SP4_COVER']] = transect_basic[
        ['LITTER_ADJUSTED', 'EXPOSED_GROUND_ADJUSTED', 'TOTAL_VEG_ADJUSTED', 'PG_TOTAL_ADJUSTED', 'AG_TOTAL_ADJUSTED',
         'PF_TOTAL_ADJUSTED', 'AF_TOTAL_ADJUSTED', 'UP_TOTAL_ADJUSTED', 'PG_3P_SP1_COVER', 'PG_3P_SP2_COVER',
         'PG_3P_SP3_COVER', 'PG_3P_SP4_COVER', 'PG_OTHER_SP1_COVER', 'PG_OTHER_SP2_COVER', 'PG_OTHER_SP3_COVER',
         'PG_OTHER_SP4_COVER', 'AG_SP1_COVER', 'AG_SP2_COVER', 'AG_SP3_COVER', 'AG_SP4_COVER', 'PF_SP1_COVER',
         'PF_SP2_COVER', 'PF_SP3_COVER', 'PF_SP4_COVER', 'AF_SP1_COVER', 'AF_SP2_COVER', 'AF_SP3_COVER',
         'AF_SP4_COVER', 'UP_SP1_COVER', 'UP_SP2_COVER', 'UP_SP3_COVER', 'UP_SP4_COVER']].fillna(value=0)

    transect_info = transect_basic.fillna(value='BLANK')  # Fill remaining blanks

    # Change values in columns to required values for workbook

    officer_values = {'roojan_bista': 'Bista, Roojan', 'henry_brink': 'Brink, Henry', 'bisun_datt': 'Datt, Bisun',
                      'rowena_eastick': 'Eastick, Rowena', 'david_hooper': 'Hooper, David',
                      'eloise_kippers': 'Kippers, Eloise',
                      'daniel_mcIntyre': 'McIntyre, Daniel', 'debbie_mitchell': 'Mitchell, Debbie',
                      'john_targett': 'Targett, John',
                      'cameron_wallace': 'Wallace, Cameron', 'sam_washusen': 'Washusen, Sam'}

    direction_values = {'north': 'North', 'north_east': 'North-East', 'east': 'East', 'south_east': 'South-East',
                        'south': 'South', 'south_west': 'South-West', 'west': 'West', 'north_west': 'North-West'}

    tran_rep_values = {'yes': 'Yes', 'no': 'No'}

    transect_info['RECORDER'] = transect_info['RECORDER'].replace(officer_values)
    transect_info['ESTIMATOR'] = transect_info['ESTIMATOR'].replace(officer_values)
    transect_info['OFFSET_DIRECTION'] = transect_info['OFFSET_DIRECTION'].replace(direction_values)
    transect_info['SITE_TRAN_REP'] = transect_info['SITE_TRAN_REP'].replace(tran_rep_values)

    return transect_info


def createSiteTranData(site, transect_info):
    df2 = transect_info['SITE'] == site
    transect_site_info = transect_info[df2]

    return transect_site_info


def writeSiteEstab(path, site, main_site_final, transect_site_info):
    # Set up writer for Pandas-Excel conversion

    book = load_workbook(path + site + '.xlsx')
    writer = pd.ExcelWriter(path + site + '.xlsx', engine='openpyxl')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

    # Declare worksheet variable

    sheet = book['Step 1 - Site Establishment']

    # Create variables for the values in the CSV file

    recorder = transect_site_info.iloc[0]['RECORDER']
    estimator = transect_site_info.iloc[0]['ESTIMATOR']
    recorder_other = transect_site_info.iloc[0]['RECORDER_other']
    estimator_other = transect_site_info.iloc[0]['ESTIMATOR_other']
    property_name = main_site_final.iloc[0]['PROPERTY_NAME']
    site_id = main_site_final.iloc[0]['SITE']
    paddock_name = main_site_final.iloc[0]['PADDOCK_NAME']
    date_time = transect_site_info.iloc[0]['DATE_TIME']
    offset_direc = transect_site_info.iloc[0]['OFFSET_DIRECTION']
    offset_lat = main_site_final.iloc[0]['NOFFSET_GPS-Latitude']
    offset_lon = main_site_final.iloc[0]['NOFFSET_GPS-Longitude']
    centre_lat = main_site_final.iloc[0]['CENTRE_GPS-Latitude']
    centre_lon = main_site_final.iloc[0]['CENTRE_GPS-Longitude']
    landscape_pos = main_site_final.iloc[0]['LANDSCAPE_POS']
    soil_color = main_site_final.iloc[0]['SOIL_COLOR']
    site_desc = main_site_final.iloc[0]['SITE_DESC']
    reas_selec = main_site_final.iloc[0]['REASON_SITE']
    land_sys_name = main_site_final.iloc[0]['LAND_SYS']
    land_sys_source = main_site_final.iloc[0]['LAND_SYS_SOURCE']
    land_sys_consist = main_site_final.iloc[0]['LAND_SYS_CONSIST']
    land_sys_ifno = main_site_final.iloc[0]['ALT_LAND_SYS']
    land_sys_notes = main_site_final.iloc[0]['LAND_SYS_NOTES']
    near_water_name = main_site_final.iloc[0]['NEAR_WATER_NAME']
    near_water_type = main_site_final.iloc[0]['WATER_TYPE']
    dist_water = main_site_final.iloc[0]['DIST_NEAR_WATER']
    direc_water = main_site_final.iloc[0]['DIRECTION_WATER']
    dist_track = main_site_final.iloc[0]['DIST_TRACK']
    direc_track = main_site_final.iloc[0]['DIRECTION_TRACK']
    other_infra = main_site_final.iloc[0]['OTHER_INFRASTRUC']

    # Create variables for sheet locations and write values

    v01 = sheet['B3']
    v01.value = recorder
    v02 = sheet['B4']
    v02.value = estimator
    v03 = sheet['B5']

    # Populates the 'Any Others Present' cell. If two non-listed officers (unlikely), then recorder is used
    if recorder_other != "BLANK":
        v03.value = recorder_other
    elif estimator_other != "BLANK":
        v03.value = estimator_other
    else:
        v03.value == "BLANK"

    v04 = sheet['B6']
    v04.value = property_name
    v05 = sheet['B8']
    v05.value = site_id
    v06 = sheet['B9']
    v06.value = paddock_name
    v07 = sheet['B10']
    v07.value = date_time
    v08 = sheet['B11']
    v08.value = offset_direc
    v09 = sheet['B12']
    v09.value = "GDA94"
    v10 = sheet['B13']
    v10.value = offset_lat
    v11 = sheet['B14']
    v11.value = offset_lon
    v12 = sheet['B15']
    v12.value = centre_lat
    v13 = sheet['B16']
    v13.value = centre_lon
    v14 = sheet['B17']
    v14.value = landscape_pos
    v15 = sheet['B18']
    v15.value = soil_color
    v16 = sheet['B19']
    v16.value = site_desc
    v17 = sheet['B20']
    v17.value = reas_selec
    v18 = sheet['B21']
    v18.value = land_sys_name
    v19 = sheet['E21']
    v19.value = land_sys_source
    v20 = sheet['B22']
    v20.value = land_sys_consist
    v21 = sheet['E22']
    v21.value = land_sys_ifno
    v22 = sheet['B23']
    v22.value = land_sys_notes
    v23 = sheet['B24']
    v23.value = near_water_name
    v24 = sheet['E24']
    v24.value = near_water_type
    v25 = sheet['B25']
    v25.value = dist_water
    v26 = sheet['E25']
    v26.value = direc_water
    v27 = sheet['B26']
    v27.value = dist_track
    v28 = sheet['E26']
    v28.value = direc_track
    v29 = sheet['B27']
    v29.value = other_infra

    writer.save()


def writeVisitDetails(path, site, main_site_final, transect_site_info):
    # Set up writer for Pandas-Excel conversion

    book = load_workbook(path + site + '.xlsx')
    writer = pd.ExcelWriter(path + site + '.xlsx', engine='openpyxl')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

    # Declare worksheet variable

    sheet = book['Step 2 - Visit Details']

    # Create variables for the values in the CSV file
    # Use 'iloc[0] when there is one row of data

    recorder = transect_site_info.iloc[0]['RECORDER']
    estimator = transect_site_info.iloc[0]['ESTIMATOR']
    site_id = main_site_final.iloc[0]['SITE']
    date_time = transect_site_info.iloc[0]['DATE_TIME']
    seas_cond = main_site_final.iloc[0]['SEASON_COND']
    atm_cond = main_site_final.iloc[0]['ATM_COND']
    surf_cracks = main_site_final.iloc[0]['SURF_CRACK']
    soil_moist = main_site_final.iloc[0]['SOIL_MOIST']
    brief_site_desc = main_site_final.iloc[0]['BRIEF_SITE_DESC']

    # Create variables for sheet locations and write values

    v01 = sheet['B3']
    v01.value = recorder
    v02 = sheet['B4']
    v02.value = estimator
    v03 = sheet['B5']
    v03.value = site_id
    v04 = sheet['B6']
    v04.value = date_time
    v05 = sheet['B8']
    v05.value = seas_cond
    v06 = sheet['B9']
    v06.value = atm_cond
    v07 = sheet['B10']
    v07.value = surf_cracks
    v08 = sheet['B11']
    v08.value = soil_moist
    v09 = sheet['B12']
    v09.value = brief_site_desc

    writer.save()


def writeDisturbance(path, site, main_site_final):
    # Set up writer for Pandas-Excel conversion

    book = load_workbook(path + site + '.xlsx')
    writer = pd.ExcelWriter(path + site + '.xlsx', engine='openpyxl')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

    # Declare worksheet variable

    sheet = book['Step 3 - Disturbance Details']

    # Create variables for the values in the CSV file

    dist_clear = main_site_final.iloc[0]['CLEARING']
    dist_clear_notes = main_site_final.iloc[0]['CLEARING_NOTES']
    dist_cyc_storm = main_site_final.iloc[0]['CYCLONE_STORM']
    dist_cyc_storm_notes = main_site_final.iloc[0]['CYCLONE_STORM_NOTES']
    dieback = main_site_final.iloc[0]['DIEBACK']
    dieback_notes = main_site_final.iloc[0]['DIEBACK_NOTES']
    adj_infra = main_site_final.iloc[0]['ADJACENT_INFRASTRUCTURE']
    disturb_notes = main_site_final.iloc[0]['DIST_NOTES']
    dist_infra = main_site_final.iloc[0]['DIST_TO_INFRASTRUCTURE']
    wild_camel = main_site_final.iloc[0]['CAMEL']
    wild_rabbit = main_site_final.iloc[0]['RABBIT']
    wild_donkey = main_site_final.iloc[0]['DONKEY']
    wild_horse = main_site_final.iloc[0]['HORSE']
    wild_pig = main_site_final.iloc[0]['PIG']
    wild_buffalo = main_site_final.iloc[0]['BUFFALO']
    nat_herb = main_site_final.iloc[0]['NATIVE_HERBIVORE']
    wild_other = main_site_final.iloc[0]['OTHER']
    wild_animal_evid = main_site_final.iloc[0]['FERAL_EVID']
    fire_north_freq = main_site_final.iloc[0]['NORTH_FF']
    fire_north_int = main_site_final.iloc[0]['NORTH_FI']
    fire_south_freq = main_site_final.iloc[0]['SOUTH_FF']
    fire_south_int = main_site_final.iloc[0]['SOUTH_FI']
    cattle_pad = main_site_final.iloc[0]['CATTLE_PAD']
    trampling = main_site_final.iloc[0]['CATTLE_TRAMP']
    erodible_soil = main_site_final.iloc[0]['ERODIBLE_SOIL']
    scald_sev = main_site_final.iloc[0]['SCALDING_SEVERITY']
    scald_stab = main_site_final.iloc[0]['SCALDING_STABILITY']
    windsheet_sev = main_site_final.iloc[0]['WINDSHEETING_SEVERITY']
    windsheet_stab = main_site_final.iloc[0]['WINDSHEETING_STABILITY']
    watersheet_sev = main_site_final.iloc[0]['WATERSHEETING_SEVERITY']
    watersheet_stab = main_site_final.iloc[0]['WATERSHEETING_STABILITY']
    rill_sev = main_site_final.iloc[0]['RILLING_SEVERITY']
    rill_stab = main_site_final.iloc[0]['RILLING_STABILITY']
    gully_sev = main_site_final.iloc[0]['GULLYING_SEVERITY']
    gully_stab = main_site_final.iloc[0]['GULLYING_STABILITY']
    erosion_comments = main_site_final.iloc[0]['EROSION_COMMENTS']
    weeds = main_site_final.iloc[0]['WEEDS']
    dist_comment = main_site_final.iloc[0]['ADD_DIST_INFO']

    # Create variables for sheet locations and write values

    v01 = sheet['B3']
    v01.value = dist_clear
    v02 = sheet['D4']
    v02.value = dist_clear_notes
    v03 = sheet['B5']
    v03.value = dist_cyc_storm
    v04 = sheet['D6']
    v04.value = dist_cyc_storm_notes
    v05 = sheet['B7']
    v05.value = dieback
    v06 = sheet['D8']
    v06.value = dieback_notes
    v07 = sheet['B9']
    v07.value = adj_infra
    v08 = sheet['B10']
    v08.value = disturb_notes

    v09 = sheet['B11']
    if dist_infra == "BLANK":
        v09.value = ""
    else:
        v09.value = dist_infra

    v10 = sheet['B13']
    v10.value = wild_camel
    v11 = sheet['C13']
    v11.value = wild_rabbit
    v12 = sheet['D13']
    v12.value = wild_donkey
    v13 = sheet['E13']
    v13.value = wild_horse
    v14 = sheet['F13']
    v14.value = wild_pig
    v15 = sheet['G13']
    v15.value = wild_buffalo
    v16 = sheet['H13']
    v16.value = nat_herb
    v17 = sheet['I13']
    v17.value = wild_other
    v18 = sheet['B14']
    v18.value = wild_animal_evid
    v19 = sheet['B17']
    v19.value = fire_north_freq
    v20 = sheet['B18']
    v20.value = fire_north_int
    v21 = sheet['G17']
    v21.value = fire_south_freq
    v22 = sheet['G18']
    v22.value = fire_south_int
    v23 = sheet['D20']
    v23.value = cattle_pad
    v24 = sheet['K20']
    v24.value = trampling
    v25 = sheet['B22']
    v25.value = erodible_soil

    v26 = sheet['B24']
    if scald_sev == "BLANK":
        v26.value = 'Absent'
    else:
        v26.value = scald_sev

    v27 = sheet['C24']
    v27.value = scald_stab

    v28 = sheet['B25']
    if windsheet_sev == "BLANK":
        v28.value = 'Absent'
    else:
        v28.value = windsheet_sev

    v29 = sheet['C25']
    v29.value = windsheet_stab

    v30 = sheet['B26']
    if watersheet_sev == "BLANK":
        v30.value = 'Absent'
    else:
        v30.value = watersheet_sev

    v31 = sheet['C26']
    v31.value = watersheet_stab

    v32 = sheet['B27']
    if rill_sev == "BLANK":
        v32.value = 'Absent'
    else:
        v32.value = rill_sev

    v33 = sheet['C27']
    v33.value = rill_stab

    v34 = sheet['B28']
    if gully_sev == "BLANK":
        v34.value = 'Absent'
    else:
        v34.value = gully_sev

    v35 = sheet['C28']
    v35.value = gully_stab
    v36 = sheet['B29']
    v36.value = erosion_comments
    v37 = sheet['B30']
    v37.value = weeds
    v38 = sheet['B31']
    v38.value = dist_comment

    writer.save()


def writeSiteCondition(path, site, main_site_final):
    # Set up writer for Pandas-Excel conversion

    book = load_workbook(path + site + '.xlsx')
    writer = pd.ExcelWriter(path + site + '.xlsx', engine='openpyxl')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

    # Declare worksheet variable

    sheet = book['Step 8 - Site Condition']

    # Create variables for the values in the CSV file

    past_green = main_site_final.iloc[0]['GREENESS']
    past_green_comm = main_site_final.iloc[0]['GREENESS_COMMENT']
    past_abun = main_site_final.iloc[0]['ABUNDANCE']
    past_abun_comm = main_site_final.iloc[0]['ABUNDANCE_COMMENT']
    past_util = main_site_final.iloc[0]['PAST_UTIL_PROP']
    past_util_comm = main_site_final.iloc[0]['PAST_UTIL_COMMENT']
    past_class = main_site_final.iloc[0]['CONDITION_CLASS']
    past_score = main_site_final.iloc[0]['CONDITION_SCORE']
    visit_notes = main_site_final.iloc[0]['VISIT_NOTES']

    v01 = sheet['B3']
    v01.value = past_green
    v02 = sheet['B4']
    v02.value = past_green_comm
    v03 = sheet['B5']
    v03.value = past_abun
    v04 = sheet['B6']
    v04.value = past_abun_comm
    v05 = sheet['B7']
    v05.value = past_util
    v06 = sheet['B8']
    v06.value = past_util_comm
    v07 = sheet['B10']
    v07.value = past_score
    v08 = sheet['G10']
    v08.value = past_class
    v09 = sheet['B11']
    v09.value = visit_notes

    writer.save()


def writeCoverEstimates(path, site, transect_site_info):
    # Set up writer for Pandas-Excel conversion

    book = load_workbook(path + site + '.xlsx')
    writer = pd.ExcelWriter(path + site + '.xlsx', engine='openpyxl')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

    # Declare worksheet variable

    sheet = book['Step 7 - Cover estimates  - Tab']

    # Create variables for the values in the CSV file #

    # Variables for SITE COVER FRACTIONS
    lit_cov_est = transect_site_info.iloc[0]['LIT_SUM']
    exp_ground_cov_est = transect_site_info.iloc[0]['EXPOSED_GROUND_SUM']
    total_veg_cov_est = transect_site_info.iloc[0]['VEG_SUM_ROUND']
    lit_cov_adj = transect_site_info.iloc[0]['LITTER_ADJUSTED']
    exp_ground_cov_adj = transect_site_info.iloc[0]['EXPOSED_GROUND_ADJUSTED']
    total_veg_cov_adj = transect_site_info.iloc[0]['TOTAL_VEG_ADJUSTED']

    # Variables for Propertion of Veg Cover
    prop_veg_pg = transect_site_info.iloc[0]['PG_SUM_PROP']
    prop_veg_ag = transect_site_info.iloc[0]['AG_SUM_PROP']
    prop_veg_pf = transect_site_info.iloc[0]['PF_SUM_PROP']
    prop_veg_af = transect_site_info.iloc[0]['AF_SUM_PROP']
    prop_veg_up = transect_site_info.iloc[0]['UP_SUM_PROP']

    # Variables for Adjusted Vegetation Cover
    pg_cov_adj = transect_site_info.iloc[0]['PG_TOTAL_ADJUSTED']
    ag_cov_adj = transect_site_info.iloc[0]['AG_TOTAL_ADJUSTED']
    pf_cov_adj = transect_site_info.iloc[0]['PF_TOTAL_ADJUSTED']
    af_cov_adj = transect_site_info.iloc[0]['AF_TOTAL_ADJUSTED']
    up_cov_adj = transect_site_info.iloc[0]['UP_TOTAL_ADJUSTED']

    # Variables for VEGETATION - Perennial grasses
    pg_3P_sp1_conf_name = transect_site_info.iloc[0]['PG_3P_SP1_CONF_NAME']
    pg_3P_sp1_field_name = transect_site_info.iloc[0]['PG_3P_SP1_FIELD_NAME']
    pg_3P_sp1_cov = transect_site_info.iloc[0]['PG_3P_SP1_COVER']
    pg_3P_sp2_conf_name = transect_site_info.iloc[0]['PG_3P_SP2_CONF_NAME']
    pg_3P_sp2_field_name = transect_site_info.iloc[0]['PG_3P_SP2_FIELD_NAME']
    pg_3P_sp2_cov = transect_site_info.iloc[0]['PG_3P_SP2_COVER']
    pg_3P_sp3_conf_name = transect_site_info.iloc[0]['PG_3P_SP3_CONF_NAME']
    pg_3P_sp3_field_name = transect_site_info.iloc[0]['PG_3P_SP3_FIELD_NAME']
    pg_3P_sp3_cov = transect_site_info.iloc[0]['PG_3P_SP3_COVER']
    pg_3P_sp4_conf_name = transect_site_info.iloc[0]['PG_3P_SP4_CONF_NAME']
    pg_3P_sp4_field_name = transect_site_info.iloc[0]['PG_3P_SP4_FIELD_NAME']
    pg_3P_sp4_cov = transect_site_info.iloc[0]['PG_3P_SP4_COVER']
    pg_other_sp1_conf_name = transect_site_info.iloc[0]['PG_OTHER_SP1_CONF_NAME']
    pg_other_sp1_field_name = transect_site_info.iloc[0]['PG_OTHER_SP1_FIELD_NAME']
    pg_other_sp1_cov = transect_site_info.iloc[0]['PG_OTHER_SP1_COVER']
    pg_other_sp2_conf_name = transect_site_info.iloc[0]['PG_OTHER_SP2_CONF_NAME']
    pg_other_sp2_field_name = transect_site_info.iloc[0]['PG_OTHER_SP2_FIELD_NAME']
    pg_other_sp2_cov = transect_site_info.iloc[0]['PG_OTHER_SP2_COVER']
    pg_other_sp3_conf_name = transect_site_info.iloc[0]['PG_OTHER_SP3_CONF_NAME']
    pg_other_sp3_field_name = transect_site_info.iloc[0]['PG_OTHER_SP3_FIELD_NAME']
    pg_other_sp3_cov = transect_site_info.iloc[0]['PG_OTHER_SP3_COVER']
    pg_other_sp4_conf_name = transect_site_info.iloc[0]['PG_OTHER_SP4_CONF_NAME']
    pg_other_sp4_field_name = transect_site_info.iloc[0]['PG_OTHER_SP4_FIELD_NAME']
    pg_other_sp4_cov = transect_site_info.iloc[0]['PG_OTHER_SP4_COVER']

    # Variables for VEGETATION - Annual grasses
    ag_sp1_conf_name = transect_site_info.iloc[0]['AG_SP1_CONF_NAME']
    ag_sp1_field_name = transect_site_info.iloc[0]['AG_SP1_FIELD_NAME']
    ag_sp1_cov = transect_site_info.iloc[0]['AG_SP1_COVER']
    ag_sp2_conf_name = transect_site_info.iloc[0]['AG_SP2_CONF_NAME']
    ag_sp2_field_name = transect_site_info.iloc[0]['AG_SP2_FIELD_NAME']
    ag_sp2_cov = transect_site_info.iloc[0]['AG_SP2_COVER']
    ag_sp3_conf_name = transect_site_info.iloc[0]['AG_SP3_CONF_NAME']
    ag_sp3_field_name = transect_site_info.iloc[0]['AG_SP3_FIELD_NAME']
    ag_sp3_cov = transect_site_info.iloc[0]['AG_SP3_COVER']
    ag_sp4_conf_name = transect_site_info.iloc[0]['AG_SP4_CONF_NAME']
    ag_sp4_field_name = transect_site_info.iloc[0]['AG_SP4_FIELD_NAME']
    ag_sp4_cov = transect_site_info.iloc[0]['AG_SP4_COVER']

    # Variables for VEGETATION - Perennial forbs
    pf_sp1_conf_name = transect_site_info.iloc[0]['PF_SP1_CONF_NAME']
    pf_sp1_field_name = transect_site_info.iloc[0]['PF_SP1_FIELD_NAME']
    pf_sp1_cov = transect_site_info.iloc[0]['PF_SP1_COVER']
    pf_sp2_conf_name = transect_site_info.iloc[0]['PF_SP2_CONF_NAME']
    pf_sp2_field_name = transect_site_info.iloc[0]['PF_SP2_FIELD_NAME']
    pf_sp2_cov = transect_site_info.iloc[0]['PF_SP2_COVER']
    pf_sp3_conf_name = transect_site_info.iloc[0]['PF_SP3_CONF_NAME']
    pf_sp3_field_name = transect_site_info.iloc[0]['PF_SP3_FIELD_NAME']
    pf_sp3_cov = transect_site_info.iloc[0]['PF_SP3_COVER']
    pf_sp4_conf_name = transect_site_info.iloc[0]['PF_SP4_CONF_NAME']
    pf_sp4_field_name = transect_site_info.iloc[0]['PF_SP4_FIELD_NAME']
    pf_sp4_cov = transect_site_info.iloc[0]['PF_SP4_COVER']

    # Variables for VEGETATION - Annual forbs
    af_sp1_conf_name = transect_site_info.iloc[0]['AF_SP1_CONF_NAME']
    af_sp1_field_name = transect_site_info.iloc[0]['AF_SP1_FIELD_NAME']
    af_sp1_cov = transect_site_info.iloc[0]['AF_SP1_COVER']
    af_sp2_conf_name = transect_site_info.iloc[0]['AF_SP2_CONF_NAME']
    af_sp2_field_name = transect_site_info.iloc[0]['AF_SP2_FIELD_NAME']
    af_sp2_cov = transect_site_info.iloc[0]['AF_SP2_COVER']
    af_sp3_conf_name = transect_site_info.iloc[0]['AF_SP3_CONF_NAME']
    af_sp3_field_name = transect_site_info.iloc[0]['AF_SP3_FIELD_NAME']
    af_sp3_cov = transect_site_info.iloc[0]['AF_SP3_COVER']
    af_sp4_conf_name = transect_site_info.iloc[0]['AF_SP4_CONF_NAME']
    af_sp4_field_name = transect_site_info.iloc[0]['AF_SP4_FIELD_NAME']
    af_sp4_cov = transect_site_info.iloc[0]['AF_SP4_COVER']

    # Variables for VEGETATION - Unlisted plants
    up_sp1_conf_name = transect_site_info.iloc[0]['UP_SP1_CONF_NAME']
    up_sp1_field_name = transect_site_info.iloc[0]['UP_SP1_FIELD_NAME']
    up_sp1_cov = transect_site_info.iloc[0]['UP_SP1_COVER']
    up_sp2_conf_name = transect_site_info.iloc[0]['UP_SP2_CONF_NAME']
    up_sp2_field_name = transect_site_info.iloc[0]['UP_SP2_FIELD_NAME']
    up_sp2_cov = transect_site_info.iloc[0]['UP_SP2_COVER']
    up_sp3_conf_name = transect_site_info.iloc[0]['UP_SP3_CONF_NAME']
    up_sp3_field_name = transect_site_info.iloc[0]['UP_SP3_FIELD_NAME']
    up_sp3_cov = transect_site_info.iloc[0]['UP_SP3_COVER']
    up_sp4_conf_name = transect_site_info.iloc[0]['UP_SP4_CONF_NAME']
    up_sp4_field_name = transect_site_info.iloc[0]['UP_SP4_FIELD_NAME']
    up_sp4_cov = transect_site_info.iloc[0]['UP_SP4_COVER']

    # Create variables for sheet locations and write values #

    # Site cover fractions
    v01 = sheet['B4']
    v01.value = lit_cov_est
    v02 = sheet['C4']
    v02.value = exp_ground_cov_est
    v03 = sheet['D4']
    v03.value = total_veg_cov_est
    v04 = sheet['B5']
    v04.value = lit_cov_adj
    v05 = sheet['C5']
    v05.value = exp_ground_cov_adj
    v06 = sheet['D5']
    v06.value = total_veg_cov_adj

    # Vegetation cover fractions
    v09 = sheet['B8']
    v09.value = prop_veg_pg
    v10 = sheet['C8']
    v10.value = prop_veg_ag
    v11 = sheet['D8']
    v11.value = prop_veg_pf
    v12 = sheet['E8']
    v12.value = prop_veg_af
    v13 = sheet['F8']
    v13.value = prop_veg_up
    v14 = sheet['B9']
    v14.value = pg_cov_adj
    v15 = sheet['C9']
    v15.value = ag_cov_adj
    v16 = sheet['D9']
    v16.value = pf_cov_adj
    v17 = sheet['E9']
    v17.value = af_cov_adj
    v18 = sheet['F9']
    v18.value = up_cov_adj

    # 3P grasses
    v20 = sheet['C16']
    v20.value = pg_3P_sp1_conf_name
    v21 = sheet['D16']
    v21.value = pg_3P_sp1_field_name
    v22 = sheet['E16']
    v22.value = pg_3P_sp1_cov
    v23 = sheet['C17']
    v23.value = pg_3P_sp2_conf_name
    v24 = sheet['D17']
    v24.value = pg_3P_sp2_field_name
    v25 = sheet['E17']
    v25.value = pg_3P_sp2_cov
    v26 = sheet['C18']
    v26.value = pg_3P_sp3_conf_name
    v27 = sheet['D18']
    v27.value = pg_3P_sp3_field_name
    v28 = sheet['E18']
    v28.value = pg_3P_sp3_cov
    v29 = sheet['C19']
    v29.value = pg_3P_sp4_conf_name
    v30 = sheet['D19']
    v30.value = pg_3P_sp4_field_name
    v31 = sheet['E19']
    v31.value = pg_3P_sp4_cov

    # other perennial grasses
    v32 = sheet['C21']
    v32.value = pg_other_sp1_conf_name
    v33 = sheet['D21']
    v33.value = pg_other_sp1_field_name
    v34 = sheet['E21']
    v34.value = pg_other_sp1_cov
    v35 = sheet['C22']
    v35.value = pg_other_sp2_conf_name
    v36 = sheet['D22']
    v36.value = pg_other_sp2_field_name
    v37 = sheet['E22']
    v37.value = pg_other_sp2_cov
    v38 = sheet['C23']
    v38.value = pg_other_sp3_conf_name
    v39 = sheet['D23']
    v39.value = pg_other_sp3_field_name
    v40 = sheet['E23']
    v40.value = pg_other_sp3_cov
    v41 = sheet['C24']
    v41.value = pg_other_sp4_conf_name
    v42 = sheet['D24']
    v42.value = pg_other_sp4_field_name
    v43 = sheet['E24']
    v43.value = pg_other_sp4_cov

    # annual grasses
    v44 = sheet['C28']
    v44.value = ag_sp1_conf_name
    v45 = sheet['D28']
    v45.value = ag_sp1_field_name
    v46 = sheet['E28']
    v46.value = ag_sp1_cov
    v47 = sheet['C29']
    v47.value = ag_sp2_conf_name
    v48 = sheet['D29']
    v48.value = ag_sp2_field_name
    v49 = sheet['E29']
    v49.value = ag_sp2_cov
    v50 = sheet['C30']
    v50.value = ag_sp3_conf_name
    v51 = sheet['D30']
    v51.value = ag_sp3_field_name
    v52 = sheet['E30']
    v52.value = ag_sp3_cov
    v53 = sheet['C31']
    v53.value = ag_sp4_conf_name
    v54 = sheet['D31']
    v54.value = ag_sp4_field_name
    v55 = sheet['E31']
    v55.value = ag_sp4_cov

    # perennial forbs
    v56 = sheet['C35']
    v56.value = pf_sp1_conf_name
    v57 = sheet['D35']
    v57.value = pf_sp1_field_name
    v58 = sheet['E35']
    v58.value = pf_sp1_cov
    v59 = sheet['C36']
    v59.value = pf_sp2_conf_name
    v60 = sheet['D36']
    v60.value = pf_sp2_field_name
    v61 = sheet['E36']
    v61.value = pf_sp2_cov
    v62 = sheet['C37']
    v62.value = pf_sp3_conf_name
    v63 = sheet['D37']
    v63.value = pf_sp3_field_name
    v64 = sheet['E37']
    v64.value = pf_sp3_cov
    v65 = sheet['C38']
    v65.value = pf_sp4_conf_name
    v66 = sheet['D38']
    v66.value = pf_sp4_field_name
    v67 = sheet['E38']
    v67.value = pf_sp4_cov

    # annual forbs
    v68 = sheet['C42']
    v68.value = af_sp1_conf_name
    v69 = sheet['D42']
    v69.value = af_sp1_field_name
    v70 = sheet['E42']
    v70.value = af_sp1_cov
    v71 = sheet['C43']
    v71.value = af_sp2_conf_name
    v72 = sheet['D43']
    v72.value = af_sp2_field_name
    v73 = sheet['E43']
    v73.value = af_sp2_cov
    v74 = sheet['C44']
    v74.value = af_sp3_conf_name
    v75 = sheet['D44']
    v75.value = af_sp3_field_name
    v76 = sheet['E44']
    v76.value = af_sp3_cov
    v77 = sheet['C45']
    v77.value = af_sp4_conf_name
    v78 = sheet['D45']
    v78.value = af_sp4_field_name
    v79 = sheet['E45']
    v79.value = af_sp4_cov

    # unlisted plants
    v80 = sheet['C49']
    v80.value = up_sp1_conf_name
    v81 = sheet['D49']
    v81.value = up_sp1_field_name
    v82 = sheet['E49']
    v82.value = up_sp1_cov
    v83 = sheet['C50']
    v83.value = up_sp2_conf_name
    v84 = sheet['D50']
    v84.value = up_sp2_field_name
    v85 = sheet['E50']
    v85.value = up_sp2_cov
    v86 = sheet['C51']
    v86.value = up_sp3_conf_name
    v87 = sheet['D51']
    v87.value = up_sp3_field_name
    v88 = sheet['E51']
    v88.value = up_sp3_cov
    v89 = sheet['C52']
    v89.value = up_sp4_conf_name
    v90 = sheet['D52']
    v90.value = up_sp4_field_name
    v91 = sheet['E52']
    v91.value = up_sp4_cov

    # Proportion of veg cover
    v92 = sheet['A15']
    v92.value = prop_veg_pg
    v93 = sheet['A28']
    v93.value = prop_veg_ag
    v94 = sheet['A35']
    v94.value = prop_veg_pf
    v95 = sheet['A42']
    v95.value = prop_veg_af
    v96 = sheet['A49']
    v96.value = prop_veg_up

    # Formulas for calculating cover totals
    sheet['E4'] = '= SUM( B4:D4 )'  # Site cover fraction - transect
    sheet['E5'] = '= SUM( B5:D5 )'  # Site cover fraction - adjusted
    sheet['B10'] = '= SUM( B9:F9 )'  # Vegetation cover - adjusted
    sheet['E25'] = '=ROUND( A15-SUM( E16:E24 ), 0 )'  # Other species - perennial grasses
    sheet['E32'] = '=ROUND( A28-SUM( E28:E31 ), 0 )'  # Other species - annual grasses
    sheet['E39'] = '=ROUND( A35-SUM( E35:E38 ), 0 )'  # Other species - perennial forbs
    sheet['E46'] = '=ROUND( A42-SUM( E42:E45 ), 0 )'  # Other species - annual forbs
    sheet['E53'] = '=ROUND( A49-SUM( E49:E52 ), 0 )'  # Other species - unspecified plants

    writer.save()


def writeGroundLayer(path, site, transect_site_info):
    # Set up writer for Pandas-Excel conversion

    book = load_workbook(path + site + '.xlsx')
    writer = pd.ExcelWriter(path + site + '.xlsx', engine='openpyxl')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

    # Declare worksheet variable

    sheet = book['Output 1 - Ground layer composi']

    # Create variables for the values in the CSV file

    transect_rep_site = transect_site_info.iloc[0]['SITE_TRAN_REP']
    pg_site_prop = transect_site_info.iloc[0]['PG_SUM']
    pg_cov_est = transect_site_info.iloc[0]['PG_SUM_PROP']
    ag_site_prop = transect_site_info.iloc[0]['AG_SUM']
    ag_cov_est = transect_site_info.iloc[0]['AG_SUM_PROP']
    pf_site_prop = transect_site_info.iloc[0]['PF_SUM']
    pf_cov_est = transect_site_info.iloc[0]['PF_SUM_PROP']
    af_site_prop = transect_site_info.iloc[0]['AF_SUM']
    af_cov_est = transect_site_info.iloc[0]['AF_SUM_PROP']
    up_site_prop = transect_site_info.iloc[0]['UP_SUM']
    up_cov_est = transect_site_info.iloc[0]['UP_SUM_PROP']
    total_veg_cov_est = transect_site_info.iloc[0]['VEG_SUM']
    lit_cov_est = transect_site_info.iloc[0]['LIT_SUM']
    exp_ground_cov_est = transect_site_info.iloc[0]['EXPOSED_GROUND_SUM']

    # Create variables for sheet locations and write values

    v1 = sheet['C3']
    v1.value = transect_rep_site
    v2 = sheet['B6']
    v2.value = pg_site_prop
    v3 = sheet['C6']
    v3.value = pg_cov_est
    v4 = sheet['B7']
    v4.value = ag_site_prop
    v5 = sheet['C7']
    v5.value = ag_cov_est
    v6 = sheet['B8']
    v6.value = pf_site_prop
    v7 = sheet['C8']
    v7.value = pf_cov_est
    v8 = sheet['B9']
    v8.value = af_site_prop
    v9 = sheet['C9']
    v9.value = af_cov_est
    v10 = sheet['B10']
    v10.value = up_site_prop
    v11 = sheet['C10']
    v11.value = up_cov_est
    v12 = sheet['B11']
    v12.value = total_veg_cov_est
    v13 = sheet['B12']
    v13.value = lit_cov_est
    v14 = sheet['B13']
    v14.value = exp_ground_cov_est
    v15 = sheet['D6']
    v15.value = pg_site_prop
    v16 = sheet['E6']
    v16.value = pg_cov_est
    v17 = sheet['D7']
    v17.value = ag_site_prop
    v18 = sheet['E7']
    v18.value = ag_cov_est
    v19 = sheet['D8']
    v19.value = pf_site_prop
    v20 = sheet['E8']
    v20.value = pf_cov_est
    v21 = sheet['D9']
    v21.value = af_site_prop
    v22 = sheet['E9']
    v22.value = af_cov_est
    v23 = sheet['D10']
    v23.value = up_site_prop
    v24 = sheet['E10']
    v24.value = up_cov_est
    v25 = sheet['D11']
    v25.value = total_veg_cov_est
    v26 = sheet['D12']
    v26.value = lit_cov_est
    v27 = sheet['D13']
    v27.value = exp_ground_cov_est

    writer.save()


def createBasalInfo(site, basal_basic):
    basal_slice = basal_basic['SITE'] == site
    basal_df = basal_basic[basal_slice]
    basal_info = basal_df.fillna(value='BLANK')

    # Change values in columns to required values for workbook
    func_type_values = {'tree': 'Tree', 'shrub': 'Shrub'}
    location_values = {'NORTH_25M': 'North', 'CENTRE_50M': 'Centre', 'SOUTH_75M': 'South',
                       'SOUTH_EAST_25M': 'South East',
                       'NORTH_WEST_75M': 'North West', 'NORTH_EAST_75M': 'North East', 'SOUTH_WEST_25M': 'South West'}
    record_basal_values = {'yes': 'Yes', 'no': 'No'}

    basal_info['RECORD_BASAL'] = basal_info['RECORD_BASAL'].replace(record_basal_values)

    basal_info['WS1_FUNC_TYPE'] = basal_info['WS1_FUNC_TYPE'].replace(func_type_values)
    basal_info['WS2_FUNC_TYPE'] = basal_info['WS2_FUNC_TYPE'].replace(func_type_values)
    basal_info['WS3_FUNC_TYPE'] = basal_info['WS3_FUNC_TYPE'].replace(func_type_values)
    basal_info['WS4_FUNC_TYPE'] = basal_info['WS4_FUNC_TYPE'].replace(func_type_values)
    basal_info['WS5_FUNC_TYPE'] = basal_info['WS5_FUNC_TYPE'].replace(func_type_values)
    basal_info['WS6_FUNC_TYPE'] = basal_info['WS6_FUNC_TYPE'].replace(func_type_values)
    basal_info['WS7_FUNC_TYPE'] = basal_info['WS7_FUNC_TYPE'].replace(func_type_values)
    basal_info['WS8_FUNC_TYPE'] = basal_info['WS8_FUNC_TYPE'].replace(func_type_values)
    basal_info['WS9_FUNC_TYPE'] = basal_info['WS9_FUNC_TYPE'].replace(func_type_values)
    basal_info['WS10_FUNC_TYPE'] = basal_info['WS10_FUNC_TYPE'].replace(func_type_values)

    basal_info['LOCATION1_LABEL'] = basal_info['LOCATION1_LABEL'].replace(location_values)
    basal_info['LOCATION2_LABEL'] = basal_info['LOCATION2_LABEL'].replace(location_values)
    basal_info['LOCATION3_LABEL'] = basal_info['LOCATION3_LABEL'].replace(location_values)
    basal_info['LOCATION4_LABEL'] = basal_info['LOCATION4_LABEL'].replace(location_values)
    basal_info['LOCATION5_LABEL'] = basal_info['LOCATION5_LABEL'].replace(location_values)
    basal_info['LOCATION6_LABEL'] = basal_info['LOCATION6_LABEL'].replace(location_values)
    basal_info['LOCATION7_LABEL'] = basal_info['LOCATION7_LABEL'].replace(location_values)

    return basal_info


def writeBasalArea(path, site, basal_info):
    # Set up writer for Pandas-Excel conversion

    book = load_workbook(path + site + '.xlsx')
    writer = pd.ExcelWriter(path + site + '.xlsx', engine='openpyxl')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

    # Declare worksheet variable

    sheet = book['Step 5 - Basal Sweeps - Table 2']

    # Create variables for the values in the CSV file #

    # Does site have recordable basal area?
    basal_record = basal_info.iloc[0]['RECORD_BASAL']

    # Counts

    loc1_name = basal_info.iloc[0]['LOCATION1_LABEL']
    loc1_factor = basal_info.iloc[0]['LOCATION1_FACTOR']
    loc1_live_tree = basal_info.iloc[0]['LOCATION1_TREE_LIVE']
    loc1_dead_tree = basal_info.iloc[0]['LOCATION1_TREE_DEAD']
    loc1_live_shrub = basal_info.iloc[0]['LOCATION1_SHRUB_LIVE']
    loc1_dead_shrub = basal_info.iloc[0]['LOCATION1_SHRUB_DEAD']

    loc2_name = basal_info.iloc[0]['LOCATION2_LABEL']
    loc2_factor = basal_info.iloc[0]['LOCATION2_FACTOR']
    loc2_live_tree = basal_info.iloc[0]['LOCATION2_TREE_LIVE']
    loc2_dead_tree = basal_info.iloc[0]['LOCATION2_TREE_DEAD']
    loc2_live_shrub = basal_info.iloc[0]['LOCATION2_SHRUB_LIVE']
    loc2_dead_shrub = basal_info.iloc[0]['LOCATION2_SHRUB_DEAD']

    loc3_name = basal_info.iloc[0]['LOCATION3_LABEL']
    loc3_factor = basal_info.iloc[0]['LOCATION3_FACTOR']
    loc3_live_tree = basal_info.iloc[0]['LOCATION3_TREE_LIVE']
    loc3_dead_tree = basal_info.iloc[0]['LOCATION3_TREE_DEAD']
    loc3_live_shrub = basal_info.iloc[0]['LOCATION3_SHRUB_LIVE']
    loc3_dead_shrub = basal_info.iloc[0]['LOCATION3_SHRUB_DEAD']

    loc4_name = basal_info.iloc[0]['LOCATION4_LABEL']
    loc4_factor = basal_info.iloc[0]['LOCATION4_FACTOR']
    loc4_live_tree = basal_info.iloc[0]['LOCATION4_TREE_LIVE']
    loc4_dead_tree = basal_info.iloc[0]['LOCATION4_TREE_DEAD']
    loc4_live_shrub = basal_info.iloc[0]['LOCATION4_SHRUB_LIVE']
    loc4_dead_shrub = basal_info.iloc[0]['LOCATION4_SHRUB_DEAD']

    loc5_name = basal_info.iloc[0]['LOCATION5_LABEL']
    loc5_factor = basal_info.iloc[0]['LOCATION5_FACTOR']
    loc5_live_tree = basal_info.iloc[0]['LOCATION5_TREE_LIVE']
    loc5_dead_tree = basal_info.iloc[0]['LOCATION5_TREE_DEAD']
    loc5_live_shrub = basal_info.iloc[0]['LOCATION5_SHRUB_LIVE']
    loc5_dead_shrub = basal_info.iloc[0]['LOCATION5_SHRUB_DEAD']

    loc6_name = basal_info.iloc[0]['LOCATION6_LABEL']
    loc6_factor = basal_info.iloc[0]['LOCATION6_FACTOR']
    loc6_live_tree = basal_info.iloc[0]['LOCATION6_TREE_LIVE']
    loc6_dead_tree = basal_info.iloc[0]['LOCATION6_TREE_DEAD']
    loc6_live_shrub = basal_info.iloc[0]['LOCATION6_SHRUB_LIVE']
    loc6_dead_shrub = basal_info.iloc[0]['LOCATION6_SHRUB_DEAD']

    loc7_name = basal_info.iloc[0]['LOCATION7_LABEL']
    loc7_factor = basal_info.iloc[0]['LOCATION7_FACTOR']
    loc7_live_tree = basal_info.iloc[0]['LOCATION7_TREE_LIVE']
    loc7_dead_tree = basal_info.iloc[0]['LOCATION7_TREE_DEAD']
    loc7_live_shrub = basal_info.iloc[0]['LOCATION7_SHRUB_LIVE']
    loc7_dead_shrub = basal_info.iloc[0]['LOCATION7_SHRUB_DEAD']

    # Totals

    at_basal_area = basal_info.iloc[0]['BA_ADULT_TREES_ROUND']
    as_basal_area = basal_info.iloc[0]['BA_ADULT_SHRUBS_ROUND']
    total_basal_area = basal_info.iloc[0]['BA_TOTAL']

    # Woody species

    ws1_conf_name = basal_info.iloc[0]['WS1_CONF_NAME']
    ws1_field_name = basal_info.iloc[0]['WS1_FIELD_NAME']
    ws1_func_type = basal_info.iloc[0]['WS1_FUNC_TYPE']

    ws2_conf_name = basal_info.iloc[0]['WS2_CONF_NAME']
    ws2_field_name = basal_info.iloc[0]['WS2_FIELD_NAME']
    ws2_func_type = basal_info.iloc[0]['WS2_FUNC_TYPE']

    ws3_conf_name = basal_info.iloc[0]['WS3_CONF_NAME']
    ws3_field_name = basal_info.iloc[0]['WS3_FIELD_NAME']
    ws3_func_type = basal_info.iloc[0]['WS3_FUNC_TYPE']

    ws4_conf_name = basal_info.iloc[0]['WS4_CONF_NAME']
    ws4_field_name = basal_info.iloc[0]['WS4_FIELD_NAME']
    ws4_func_type = basal_info.iloc[0]['WS4_FUNC_TYPE']

    ws5_conf_name = basal_info.iloc[0]['WS5_CONF_NAME']
    ws5_field_name = basal_info.iloc[0]['WS5_FIELD_NAME']
    ws5_func_type = basal_info.iloc[0]['WS5_FUNC_TYPE']

    ws6_conf_name = basal_info.iloc[0]['WS6_CONF_NAME']
    ws6_field_name = basal_info.iloc[0]['WS6_FIELD_NAME']
    ws6_func_type = basal_info.iloc[0]['WS6_FUNC_TYPE']

    ws7_conf_name = basal_info.iloc[0]['WS7_CONF_NAME']
    ws7_field_name = basal_info.iloc[0]['WS7_FIELD_NAME']
    ws7_func_type = basal_info.iloc[0]['WS7_FUNC_TYPE']

    ws8_conf_name = basal_info.iloc[0]['WS8_CONF_NAME']
    ws8_field_name = basal_info.iloc[0]['WS8_FIELD_NAME']
    ws8_func_type = basal_info.iloc[0]['WS8_FUNC_TYPE']

    ws9_conf_name = basal_info.iloc[0]['WS9_CONF_NAME']
    ws9_field_name = basal_info.iloc[0]['WS9_FIELD_NAME']
    ws9_func_type = basal_info.iloc[0]['WS9_FUNC_TYPE']

    ws10_conf_name = basal_info.iloc[0]['WS10_CONF_NAME']
    ws10_field_name = basal_info.iloc[0]['WS10_FIELD_NAME']
    ws10_func_type = basal_info.iloc[0]['WS10_FUNC_TYPE']

    ## Create variables for sheet locations and write values##

    # Individual basal values
    v01 = sheet['B3']
    v01.value = loc1_name
    v02 = sheet['D3']
    v02.value = loc2_name
    v03 = sheet['F3']
    v03.value = loc3_name
    v04 = sheet['H3']
    v04.value = loc4_name
    v05 = sheet['J3']
    v05.value = loc5_name
    v06 = sheet['L3']
    v06.value = loc6_name
    v07 = sheet['N3']
    v07.value = loc7_name
    v08 = sheet['B4']
    v08.value = loc1_factor
    v09 = sheet['D4']
    v09.value = loc2_factor
    v10 = sheet['F4']
    v10.value = loc3_factor
    v11 = sheet['H4']
    v11.value = loc4_factor
    v12 = sheet['J4']
    v12.value = loc5_factor
    v13 = sheet['L4']
    v13.value = loc6_factor
    v14 = sheet['N4']
    v14.value = loc7_factor
    v15 = sheet['B7']
    v15.value = loc1_live_tree
    v16 = sheet['C7']
    v16.value = loc1_dead_tree
    v17 = sheet['B8']
    v17.value = loc1_live_shrub
    v18 = sheet['C8']
    v18.value = loc1_dead_shrub
    v19 = sheet['D7']
    v19.value = loc2_live_tree
    v20 = sheet['E7']
    v20.value = loc2_dead_tree
    v21 = sheet['D8']
    v21.value = loc2_live_shrub
    v22 = sheet['E8']
    v22.value = loc2_dead_shrub
    v23 = sheet['F7']
    v23.value = loc3_live_tree
    v24 = sheet['G7']
    v24.value = loc3_dead_tree
    v25 = sheet['F8']
    v25.value = loc3_live_shrub
    v26 = sheet['G8']
    v26.value = loc3_dead_shrub
    v27 = sheet['H7']
    v27.value = loc4_live_tree
    v28 = sheet['I7']
    v28.value = loc4_dead_tree
    v29 = sheet['H8']
    v29.value = loc4_live_shrub
    v30 = sheet['I8']
    v30.value = loc4_dead_shrub
    v31 = sheet['J7']
    v31.value = loc5_live_tree
    v32 = sheet['K7']
    v32.value = loc5_dead_tree
    v33 = sheet['J8']
    v33.value = loc5_live_shrub
    v34 = sheet['K8']
    v34.value = loc5_dead_shrub
    v35 = sheet['L7']
    v35.value = loc6_live_tree
    v36 = sheet['M7']
    v36.value = loc6_dead_tree
    v37 = sheet['L8']
    v37.value = loc6_live_shrub
    v38 = sheet['M8']
    v38.value = loc6_dead_shrub
    v39 = sheet['N7']
    v39.value = loc7_live_tree
    v40 = sheet['O7']
    v40.value = loc7_dead_tree
    v41 = sheet['N8']
    v41.value = loc7_live_shrub
    v42 = sheet['O8']
    v42.value = loc7_dead_shrub

    # Overall basal values
    v43 = sheet['B17']
    v43.value = at_basal_area
    v44 = sheet['B18']
    v44.value = as_basal_area
    v45 = sheet['B19']
    v45.value = total_basal_area

    # Woody species
    v46 = sheet['A23']
    v46.value = ws1_conf_name
    v47 = sheet['F23']
    v47.value = ws1_field_name
    v48 = sheet['L23']
    v48.value = ws1_func_type
    v49 = sheet['A24']
    v49.value = ws2_conf_name
    v50 = sheet['F24']
    v50.value = ws2_field_name
    v51 = sheet['L24']
    v51.value = ws2_func_type
    v52 = sheet['A25']
    v52.value = ws3_conf_name
    v53 = sheet['F25']
    v53.value = ws3_field_name
    v54 = sheet['L25']
    v54.value = ws3_func_type
    v55 = sheet['A26']
    v55.value = ws4_conf_name
    v56 = sheet['F26']
    v56.value = ws4_field_name
    v57 = sheet['L26']
    v57.value = ws4_func_type
    v58 = sheet['A27']
    v58.value = ws5_conf_name
    v59 = sheet['F27']
    v59.value = ws5_field_name
    v60 = sheet['L27']
    v60.value = ws5_func_type
    v61 = sheet['A28']
    v61.value = ws6_conf_name
    v62 = sheet['F28']
    v62.value = ws6_field_name
    v63 = sheet['L28']
    v63.value = ws6_func_type
    v64 = sheet['A29']
    v64.value = ws7_conf_name
    v65 = sheet['F29']
    v65.value = ws7_field_name
    v66 = sheet['L29']
    v66.value = ws7_func_type
    v67 = sheet['A30']
    v67.value = ws8_conf_name
    v68 = sheet['F30']
    v68.value = ws8_field_name
    v69 = sheet['L30']
    v69.value = ws8_func_type
    v70 = sheet['A31']
    v70.value = ws9_conf_name
    v71 = sheet['F31']
    v71.value = ws9_field_name
    v72 = sheet['L31']
    v72.value = ws9_func_type
    v73 = sheet['A32']
    v73.value = ws10_conf_name
    v74 = sheet['F32']
    v74.value = ws10_field_name
    v75 = sheet['L32']
    v75.value = ws10_func_type
    v76 = sheet['L2']
    v76.value = basal_record

    writer.save()


def createWoodyInfo(site, woody_basic):
    woody_slice = woody_basic['SITE'] == site
    woody_df = woody_basic[woody_slice]
    woody_info = woody_df.fillna(value='BLANK')

    # Change values in columns to required values for workbook

    question_values = {'yes': 'Yes', 'no': 'NO'}
    belt_width_values = {'500': '1 metre', '1000': '2 metres', '2000': '4 metres'}
    func_type_values = {'tree': 'Tree', 'shrub': 'Shrub'}

    woody_info['WOODY_QUESTION'] = woody_info['WOODY_QUESTION'].replace(question_values)
    woody_info['BELT_WIDTH'] = woody_info['BELT_WIDTH'].replace(belt_width_values)
    woody_info['WS1_FUNC_TYPE'] = woody_info['WS1_FUNC_TYPE'].replace(func_type_values)
    woody_info['WS2_FUNC_TYPE'] = woody_info['WS2_FUNC_TYPE'].replace(func_type_values)
    woody_info['WS3_FUNC_TYPE'] = woody_info['WS3_FUNC_TYPE'].replace(func_type_values)
    woody_info['WS4_FUNC_TYPE'] = woody_info['WS4_FUNC_TYPE'].replace(func_type_values)
    woody_info['WS5_FUNC_TYPE'] = woody_info['WS5_FUNC_TYPE'].replace(func_type_values)
    woody_info['WS6_FUNC_TYPE'] = woody_info['WS6_FUNC_TYPE'].replace(func_type_values)
    woody_info['WS7_FUNC_TYPE'] = woody_info['WS7_FUNC_TYPE'].replace(func_type_values)
    woody_info['WS8_FUNC_TYPE'] = woody_info['WS8_FUNC_TYPE'].replace(func_type_values)
    woody_info['WS9_FUNC_TYPE'] = woody_info['WS9_FUNC_TYPE'].replace(func_type_values)
    woody_info['WS10_FUNC_TYPE'] = woody_info['WS10_FUNC_TYPE'].replace(func_type_values)

    return woody_info


def writeWoodyThick(path, site, woody_info):
    # Set up writer for Pandas-Excel conversion

    book = load_workbook(path + site + '.xlsx')
    writer = pd.ExcelWriter(path + site + '.xlsx', engine='openpyxl')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

    # Declare worksheet variable

    sheet = book['STEP 6 - Juvenile stem count - ']

    question_woody = woody_info.iloc[0]['WOODY_QUESTION']
    width_belt = woody_info.iloc[0]['BELT_WIDTH']
    t1_juv_shrub = woody_info.iloc[0]['T1_JUV_SHRUB']
    t2_juv_shrub = woody_info.iloc[0]['T2_JUV_SHRUB']
    t3_juv_shrub = woody_info.iloc[0]['T3_JUV_SHRUB']
    t4_juv_shrub = woody_info.iloc[0]['T4_JUV_SHRUB']
    t5_juv_shrub = woody_info.iloc[0]['T5_JUV_SHRUB']
    t1_juv_tree = woody_info.iloc[0]['T1_JUV_TREE']
    t2_juv_tree = woody_info.iloc[0]['T2_JUV_TREE']
    t3_juv_tree = woody_info.iloc[0]['T3_JUV_TREE']
    t4_juv_tree = woody_info.iloc[0]['T4_JUV_TREE']
    t5_juv_tree = woody_info.iloc[0]['T5_JUV_TREE']
    shrub_stem_dens = woody_info.iloc[0]['STEM_DENSITY_SHRUB_CALC']
    tree_stem_dens = woody_info.iloc[0]['STEM_DENSITY_TREE_CALC']
    total_dens = woody_info.iloc[0]['TOTAL_DENSITY_CALC']
    ws1_conf_name = woody_info.iloc[0]['WS1_CONF_NAME']
    ws1_field_name = woody_info.iloc[0]['WS1_FIELD_NAME']
    ws1_func_type = woody_info.iloc[0]['WS1_FUNC_TYPE']
    ws2_conf_name = woody_info.iloc[0]['WS2_CONF_NAME']
    ws2_field_name = woody_info.iloc[0]['WS2_FIELD_NAME']
    ws2_func_type = woody_info.iloc[0]['WS2_FUNC_TYPE']
    ws3_conf_name = woody_info.iloc[0]['WS3_CONF_NAME']
    ws3_field_name = woody_info.iloc[0]['WS3_FIELD_NAME']
    ws3_func_type = woody_info.iloc[0]['WS3_FUNC_TYPE']
    ws4_conf_name = woody_info.iloc[0]['WS4_CONF_NAME']
    ws4_field_name = woody_info.iloc[0]['WS4_FIELD_NAME']
    ws4_func_type = woody_info.iloc[0]['WS4_FUNC_TYPE']
    ws5_conf_name = woody_info.iloc[0]['WS5_CONF_NAME']
    ws5_field_name = woody_info.iloc[0]['WS5_FIELD_NAME']
    ws5_func_type = woody_info.iloc[0]['WS5_FUNC_TYPE']
    ws6_conf_name = woody_info.iloc[0]['WS6_CONF_NAME']
    ws6_field_name = woody_info.iloc[0]['WS6_FIELD_NAME']
    ws6_func_type = woody_info.iloc[0]['WS6_FUNC_TYPE']
    ws7_conf_name = woody_info.iloc[0]['WS7_CONF_NAME']
    ws7_field_name = woody_info.iloc[0]['WS7_FIELD_NAME']
    ws7_func_type = woody_info.iloc[0]['WS7_FUNC_TYPE']
    ws8_conf_name = woody_info.iloc[0]['WS8_CONF_NAME']
    ws8_field_name = woody_info.iloc[0]['WS8_FIELD_NAME']
    ws8_func_type = woody_info.iloc[0]['WS8_FUNC_TYPE']
    ws9_conf_name = woody_info.iloc[0]['WS9_CONF_NAME']
    ws9_field_name = woody_info.iloc[0]['WS9_FIELD_NAME']
    ws9_func_type = woody_info.iloc[0]['WS9_FUNC_TYPE']
    ws10_conf_name = woody_info.iloc[0]['WS10_CONF_NAME']
    ws10_field_name = woody_info.iloc[0]['WS10_FIELD_NAME']
    ws10_func_type = woody_info.iloc[0]['WS10_FUNC_TYPE']

    ## Create variables for sheet locations and write values ##

    v01 = sheet['F2']
    v01.value = question_woody
    v02 = sheet['B3']
    v02.value = width_belt
    v03 = sheet['B5']
    v03.value = t1_juv_shrub
    v04 = sheet['B6']
    v04.value = t1_juv_tree
    v05 = sheet['C5']
    v05.value = t2_juv_shrub
    v06 = sheet['C6']
    v06.value = t2_juv_tree
    v07 = sheet['D5']
    v07.value = t3_juv_shrub
    v08 = sheet['D6']
    v08.value = t3_juv_tree
    v09 = sheet['E5']
    v09.value = t4_juv_shrub
    v10 = sheet['E6']
    v10.value = t4_juv_tree
    v11 = sheet['F5']
    v11.value = t5_juv_shrub
    v12 = sheet['F6']
    v12.value = t5_juv_tree
    v13 = sheet['C10']
    v13.value = shrub_stem_dens
    v14 = sheet['C11']
    v14.value = tree_stem_dens
    v15 = sheet['C12']
    v15.value = total_dens
    v16 = sheet['A15']
    v16.value = ws1_conf_name
    v17 = sheet['A16']
    v17.value = ws2_conf_name
    v18 = sheet['A17']
    v18.value = ws3_conf_name
    v19 = sheet['A18']
    v19.value = ws4_conf_name
    v20 = sheet['A19']
    v20.value = ws5_conf_name
    v21 = sheet['A20']
    v21.value = ws6_conf_name
    v22 = sheet['A21']
    v22.value = ws7_conf_name
    v23 = sheet['A22']
    v23.value = ws8_conf_name
    v24 = sheet['A23']
    v24.value = ws9_conf_name
    v25 = sheet['A24']
    v25.value = ws10_conf_name
    v26 = sheet['B15']
    v26.value = ws1_field_name
    v27 = sheet['B16']
    v27.value = ws2_field_name
    v28 = sheet['B17']
    v28.value = ws3_field_name
    v29 = sheet['B18']
    v29.value = ws4_field_name
    v30 = sheet['B19']
    v30.value = ws5_field_name
    v31 = sheet['B20']
    v31.value = ws6_field_name
    v32 = sheet['B21']
    v32.value = ws7_field_name
    v33 = sheet['B22']
    v33.value = ws8_field_name
    v34 = sheet['B23']
    v34.value = ws9_field_name
    v35 = sheet['B24']
    v35.value = ws10_field_name
    v36 = sheet['D15']
    v36.value = ws1_func_type
    v37 = sheet['D16']
    v37.value = ws2_func_type
    v38 = sheet['D17']
    v38.value = ws3_func_type
    v39 = sheet['D18']
    v39.value = ws4_func_type
    v40 = sheet['D19']
    v40.value = ws5_func_type
    v41 = sheet['D20']
    v41.value = ws6_func_type
    v42 = sheet['D21']
    v42.value = ws7_func_type
    v43 = sheet['D22']
    v43.value = ws8_func_type
    v44 = sheet['D23']
    v44.value = ws9_func_type
    v45 = sheet['D24']
    v45.value = ws10_func_type

    # Formulas for calculating transect totals
    sheet['B7'] = '= SUM( B5:B6 )'  # T1
    sheet['C7'] = '= SUM( C5:C6 )'  # T2
    sheet['D7'] = '= SUM( D5:D6 )'  # T3
    sheet['E7'] = '= SUM( E5:E6 )'  # T4
    sheet['F7'] = '= SUM( F5:F6 )'  # T5

    writer.save()


def createExcelfile(path, transect_info, main_final):
    # Output excel file that has the main fields of interest

    transect_extract = transect_info[['SITE', 'DATE_TIME', 'RECORDER', 'RECORDER_other', 'ESTIMATOR',
                                      'ESTIMATOR_other', 'OFFSET_DIRECTION', 'EXPOSED_GROUND_SUM', 'LIT_SUM', 'VEG_SUM',
                                      'PG_SUM_PROP',
                                      'AG_SUM_PROP', 'PF_SUM_PROP', 'AF_SUM_PROP', 'UP_SUM_PROP']]
    transect_extract.columns = (['SITE', 'DATE', 'RECORDER', 'RECORDER_OTHER', 'ESTIMATOR', 'ESTIMATOR_OTHER',
                                 'OFF_DIR', 'BG_%', 'LIT_%', 'VEG_%', 'PG_%', 'AG_%', 'PF_%', 'AF_%', 'UP_%'])

    main_extract = main_final[
        ['PROPERTY_NAME', 'SITE', 'GREENESS', 'ABUNDANCE', 'PAST_UTIL_PROP', 'CONDITION_CLASS', 'CONDITION_SCORE',
         'CENTRE_GPS-Latitude', 'CENTRE_GPS-Longitude', 'CENTRE_GPS-Accuracy', 'NOFFSET_GPS-Latitude',
         'NOFFSET_GPS-Longitude', 'NOFFSET_GPS-Accuracy', ]]

    main_extract.columns = (
    ['PROP', 'SITE', 'GREENESS', 'ABUNDANCE', 'UTILISATION', 'ASSES_SCORE', 'LANDCONGUIDE', 'C_LAT', 'C_LON',
     'C_ACUR', 'NO_LAT', 'NO_LON', 'NO_ACUR'])

    odk_output = pd.merge(transect_extract, main_extract, on='SITE')

    odk_output = odk_output[
        ['PROP', 'SITE', 'DATE', 'C_LAT', 'C_LON', 'C_ACUR', 'NO_LAT', 'NO_LON', 'NO_ACUR', 'RECORDER', 'ESTIMATOR',
         'RECORDER_OTHER', 'ESTIMATOR_OTHER', 'OFF_DIR', 'BG_%', 'LIT_%', 'VEG_%', 'PG_%', 'AG_%',
         'PF_%', 'AF_%', 'UP_%', 'GREENESS', 'ABUNDANCE', 'UTILISATION', 'ASSES_SCORE', 'LANDCONGUIDE']]

    cols_to_check = ['RECORDER_OTHER', 'ESTIMATOR_OTHER', 'ASSES_SCORE']
    odk_output[cols_to_check] = odk_output[cols_to_check].replace({'BLANK': ''}, regex=True)  # Remove 'BLANK' values

    odk_output.to_excel(path + 'ODK_output.xlsx')

    return odk_output


def createShapefile(path, odk_output):
    odk_output['geometry'] = [Point(xy) for xy in zip(odk_output['C_LON'], odk_output['C_LAT'])]

    odk_output_shp = odk_output[
        ['geometry', 'PROP', 'SITE', 'DATE', 'RECORDER', 'RECORDER_OTHER', 'ESTIMATOR', 'ESTIMATOR_OTHER',
         'C_LAT', 'C_LON', 'C_ACUR', 'NO_LAT', 'NO_LON', 'NO_ACUR', 'OFF_DIR', 'BG_%', 'LIT_%', 'VEG_%',
         'PG_%', 'AG_%', 'PF_%', 'AF_%', 'UP_%', 'GREENESS', 'ABUNDANCE', 'UTILISATION', 'ASSES_SCORE', 'LANDCONGUIDE']]

    crs = {'proj': 'longlat', 'ellps': 'GRS80', 'ellips': 'wgs84', 'no_defs': True}  # This method works

    # Create the geo data frame
    odk_output_shp = geopandas.GeoDataFrame(odk_output_shp, crs=crs, geometry='geometry')

    # Save out the geo data frame in esri shapefile format
    odk_output_shp.to_file(path + 'ODK_output.shp', driver='ESRI Shapefile')


def processPhotos(path, main_final, transect_info, gps_photo):
    transect_subset = transect_info[['SITE', 'DATE_TIME']]

    transect_subset.columns = (['SITE', 'DATE'])

    main_subset = main_final[['SITE', 'JPG1_CLEARING', 'JPG2_CLEARING', 'JPG3_CLEARING', 'JPG1_CYCLONE_STORM',
                              'JPG2_CYCLONE_STORM', 'JPG3_CYCLONE_STORM', 'JPG1_DIEBACK', 'JPG2_DIEBACK',
                              'JPG3_DIEBACK']]

    main_subset.columns = (
    ['SITE', 'CLEARING_01', 'CLEARING_02', 'CLEARING_03', 'CYCLONE_STORM_01', 'CYCLONE_STORM_02', 'CYCLONE_STORM_03',
     'DIEBACK_01', 'DIEBACK_02', 'DIEBACK_03'])

    gps_photos = gps_photo[['SITE', 'JPG_NO', 'JPG_C', 'JPG_N', 'JPG_NE', 'JPG_SE', 'JPG_S', 'JPG_SW', 'JPG_NW']]

    gps_photos.columns = (['SITE', 'NO', 'C', 'N', 'NE', 'SE', 'S', 'SW', 'NW'])

    df1 = pd.merge(gps_photos, main_subset, on='SITE')

    df2 = pd.merge(transect_subset, df1, on='SITE')

    merged_photo_df = df2.fillna(value='\media')

    merged_photo_df['NO'] = merged_photo_df['NO'].map(lambda x: x.lstrip('\media'))
    merged_photo_df['C'] = merged_photo_df['C'].map(lambda x: x.lstrip('\media'))
    merged_photo_df['N'] = merged_photo_df['N'].map(lambda x: x.lstrip('\media'))
    merged_photo_df['NE'] = merged_photo_df['NE'].map(lambda x: x.lstrip('\media'))
    merged_photo_df['SE'] = merged_photo_df['SE'].map(lambda x: x.lstrip('\media'))
    merged_photo_df['S'] = merged_photo_df['S'].map(lambda x: x.lstrip('\media'))
    merged_photo_df['SW'] = merged_photo_df['SW'].map(lambda x: x.lstrip('\media'))
    merged_photo_df['NW'] = merged_photo_df['NW'].map(lambda x: x.lstrip('\media'))
    merged_photo_df['CLEARING_01'] = merged_photo_df['CLEARING_01'].map(lambda x: x.lstrip('\media'))
    merged_photo_df['CLEARING_02'] = merged_photo_df['CLEARING_02'].map(lambda x: x.lstrip('\media'))
    merged_photo_df['CLEARING_03'] = merged_photo_df['CLEARING_03'].map(lambda x: x.lstrip('\media'))
    merged_photo_df['CYCLONE_STORM_01'] = merged_photo_df['CYCLONE_STORM_01'].map(lambda x: x.lstrip('\media'))
    merged_photo_df['CYCLONE_STORM_02'] = merged_photo_df['CYCLONE_STORM_02'].map(lambda x: x.lstrip('\media'))
    merged_photo_df['CYCLONE_STORM_03'] = merged_photo_df['CYCLONE_STORM_03'].map(lambda x: x.lstrip('\media'))
    merged_photo_df['DIEBACK_01'] = merged_photo_df['DIEBACK_01'].map(lambda x: x.lstrip('\media'))
    merged_photo_df['DIEBACK_02'] = merged_photo_df['DIEBACK_02'].map(lambda x: x.lstrip('\media'))
    merged_photo_df['DIEBACK_03'] = merged_photo_df['DIEBACK_03'].map(lambda x: x.lstrip('\media'))

    # Save as excel file
    merged_photo_df.to_excel(path + 'Site_photos.xlsx')


##################################         MAIN CODE       ####################################################


def mainRoutine():
    ## Set working directory for input and output files ##

    path = "G:/ODK_Field_Data_Output/"

    ########## Import CSV files ##########

    transect_basic = pd.read_csv(path + 'Rangeland Monitoring_ Star Transect.csv')
    transect1_point = pd.read_csv(path + 'Rangeland Monitoring_ Star Transect-REPEAT_points_1.csv')
    transect2_point = pd.read_csv(path + 'Rangeland Monitoring_ Star Transect-REPEAT_points_2.csv')
    transect3_point = pd.read_csv(path + 'Rangeland Monitoring_ Star Transect-REPEAT_points_3.csv')
    main_data = pd.read_csv(path + 'Rangeland Monitoring_ Integrated Site.csv')
    gps_photo = pd.read_csv(path + 'Rangeland Monitoring_ GPS Points.csv')
    basal_basic = pd.read_csv(path + 'Rangeland Monitoring_ Basal Area.csv')
    woody_basic = pd.read_csv(path + 'Rangeland Monitoring_ Woody Thickening.csv')

    ########## Prepare transect intercept data ##########

    cover_df = SplitKey(transect1_point, transect2_point, transect3_point)
    final_cover_df = PrepCover(transect_basic, cover_df)

    ########## Prepare main site data ##########

    main_final = createMainData(main_data, gps_photo)

    ########## Prepare transect data ##########

    transect_info = createTransectInfo(transect_basic)

    ########## Iterate through the cover data for each site & transect & write to excel ##########

    site_name = main_data['SITE'].unique()

    print
    "--------------------------------------"

    for site in site_name:
        CreateWorkbook(path, site)
        writeTransect(path, site, final_cover_df)
        main_site_final = createSiteMainData(site, main_final)
        transect_site_info = createSiteTranData(site, transect_info)
        writeSiteEstab(path, site, main_site_final, transect_site_info)
        writeVisitDetails(path, site, main_site_final, transect_site_info)
        writeDisturbance(path, site, main_site_final)
        writeSiteCondition(path, site, main_site_final)
        writeCoverEstimates(path, site, transect_site_info)
        writeGroundLayer(path, site, transect_site_info)
        print
        "* Site {} processed".format(site)

    for site in site_name:
        basal_info = createBasalInfo(site, basal_basic)
        if basal_info.empty:
            print
            "* No basal data for {}".format(site)
            continue
        writeBasalArea(path, site, basal_info)
        print
        "* Basal data processed for site {}".format(site)

    for site in site_name:
        woody_info = createWoodyInfo(site, woody_basic)
        if woody_info.empty:
            print
            "* No woody thickening data for {}".format(site)
            continue
        writeWoodyThick(path, site, woody_info)
        print
        "* Woody thickening data processed for site {}".format(site)

    ########## Call function to create an Excel spreadsheet for all sites ##########

    odk_output = createExcelfile(path, transect_info, main_final)
    print
    "* Excel (summary) workbook created"

    ########## Call function to generate a shapefile ##########

    createShapefile(path, odk_output)
    print
    "* Shapefile created"

    ########## Call function to extract photo columns from table and save to Excel spreadsheet ##########

    processPhotos(path, main_data, transect_basic, gps_photo)
    print
    "* Photo summary spreadsheet created"

    print
    "--------------------------------------\nPROCESSING COMPLETE!"


if __name__ == "__main__":
    mainRoutine()